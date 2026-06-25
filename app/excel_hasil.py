"""
Tulis-balik hasil pendaftaran ke Excel: kolom 'No. Tiket', 'Status Daftar',
'Waktu Daftar'. Dipakai bersama oleh tools/trial_daftar.py & tools/jalankan_batch.py
agar perilaku flag (anti-dobel) konsisten.

Aturan: baris yang 'No. Tiket'-nya sudah terisi dianggap SUDAH terdaftar dan
tidak boleh didaftarkan ulang. Saat sukses, program WAJIB menulis No. Tiket ke
baris itu (lewat tulis_hasil) supaya flag aktif.
"""
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill

KOL_TIKET = "No. Tiket"
KOL_STATUS = "Status Daftar"
KOL_WAKTU = "Waktu Daftar"

# Penanda status terminal: baris ini tidak perlu diproses ulang saat rerun.
# Dicocokkan via str(status).startswith(<prefix>). Nilai HARUS sama dengan
# `status_prefix` pada exception terkait di app/automation/ckg_bot.py.
STATUS_SUDAH_LAYANAN = "SUDAH CKG"
STATUS_TIDAK_VALID = "DATA TIDAK VALID"

# Semua penanda terminal (untuk skip saat rerun).
STATUS_TERMINAL = (STATUS_SUDAH_LAYANAN, STATUS_TIDAK_VALID)

# --- Tahap KONFIRMASI HADIR (kolom hasil terpisah dari pendaftaran) ---
KOL_STATUS_HADIR = "Status Hadir"
KOL_WAKTU_HADIR = "Waktu Hadir"
# Status saat berhasil konfirmasi kehadiran.
STATUS_HADIR = "HADIR"
# Penanda terminal hadir (skip saat rerun): sukses 'HADIR' atau portal sudah
# 'SUDAH HADIR'. Cocokkan via str(status).startswith(<prefix>).
STATUS_HADIR_TERMINAL = (STATUS_HADIR, "SUDAH HADIR")


def _cari_kolom(ws, header_row_1based, nama, buat=False):
    """Index kolom (1-based) dgn header `nama`; buat di ujung bila `buat`."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row_1based, column=c).value
        if v is not None and str(v).strip() == nama:
            return c
    if not buat:
        return None
    col = ws.max_column + 1
    ws.cell(row=header_row_1based, column=col, value=nama)
    return col


def baca_tiket(excel_path, baris_openpyxl, header_row=0):
    """Kembalikan No. Tiket pada baris itu ('' bila kosong / kolom belum ada)."""
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        col = _cari_kolom(ws, header_row + 1, KOL_TIKET, buat=False)
        if col is None:
            return ""
        v = ws.cell(row=baris_openpyxl, column=col).value
        return str(v).strip() if v is not None and str(v).strip() else ""
    except Exception:
        return ""
    finally:
        if wb is not None:
            wb.close()   # lepaskan lock file (penting di Windows)


def pastikan_kolom_hasil(excel_path, header_row=0):
    """Pastikan 3 kolom hasil ada (buat bila belum) & file bisa ditulis.
    Return (ok, pesan). ok=False bila file terkunci (masih dibuka Excel)."""
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        return False, f"gagal membuka: {type(e).__name__}: {e}"
    ws = wb.worksheets[0]
    hdr = header_row + 1
    _cari_kolom(ws, hdr, KOL_TIKET, buat=True)
    _cari_kolom(ws, hdr, KOL_STATUS, buat=True)
    _cari_kolom(ws, hdr, KOL_WAKTU, buat=True)
    try:
        wb.save(excel_path)
        return True, "ok"
    except PermissionError:
        return False, "file Excel sedang dibuka - TUTUP Excel lalu coba lagi."


# --- Pewarnaan baris menurut status (visual di Excel) -----------------------
# Warna latar (RGB hex) per kategori status. Senada gaya conditional formatting
# bawaan Excel (hijau=ok, merah=gagal, kuning=peringatan, biru=info).
WARNA_SUKSES = "C6EFCE"       # hijau muda  -> SUKSES / HADIR
WARNA_GAGAL = "FFC7CE"        # merah muda  -> GAGAL / ERROR / DATA TIDAK VALID
WARNA_PERINGATAN = "FFEB9C"   # kuning      -> DILEWATI (cek data) / peringatan
WARNA_INFO = "DDEBF7"         # biru muda   -> SUDAH CKG / SUDAH HADIR


def _kategori_status(status):
    """Petakan teks status -> kategori warna ('sukses'/'gagal'/'peringatan'/
    'info') atau None bila tidak dikenali (tidak diwarnai)."""
    s = str(status or "").strip().upper()
    if not s:
        return None
    if s.startswith("SUKSES") or s.startswith(STATUS_HADIR):
        return "sukses"
    if (s.startswith("GAGAL") or s.startswith("ERROR")
            or s.startswith(STATUS_TIDAK_VALID)):
        return "gagal"
    if s.startswith("DILEWATI") or s.startswith("PERINGATAN"):
        return "peringatan"
    if s.startswith("SUDAH"):   # SUDAH CKG / SUDAH HADIR
        return "info"
    return None


def warnai_baris(ws, baris_openpyxl, status, kolom_awal=1, kolom_akhir=None):
    """Beri warna latar SELURUH sel di baris sesuai kategori `status`.

    Tidak menyimpan workbook (pemanggil yang memanggil wb.save). Baris dengan
    status tak dikenali dibiarkan tanpa warna. Lihat `_kategori_status`.
    """
    warna = {
        "sukses": WARNA_SUKSES,
        "gagal": WARNA_GAGAL,
        "peringatan": WARNA_PERINGATAN,
        "info": WARNA_INFO,
    }.get(_kategori_status(status))
    if warna is None:
        return
    fill = PatternFill(start_color=warna, end_color=warna, fill_type="solid")
    if kolom_akhir is None:
        kolom_akhir = ws.max_column
    for c in range(kolom_awal, kolom_akhir + 1):
        ws.cell(row=baris_openpyxl, column=c).fill = fill


def tulis_hasil(excel_path, baris_openpyxl, no_tiket=None, status=None,
                header_row=0):
    """
    Tulis No. Tiket / Status Daftar / Waktu Daftar ke baris (membuat kolomnya
    bila belum ada). Kembalikan (ok, pesan); ok=False bila file terkunci.
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        return False, f"gagal membuka: {type(e).__name__}: {e}"
    ws = wb.worksheets[0]
    hdr = header_row + 1
    ct = _cari_kolom(ws, hdr, KOL_TIKET, buat=True)
    cs = _cari_kolom(ws, hdr, KOL_STATUS, buat=True)
    cw = _cari_kolom(ws, hdr, KOL_WAKTU, buat=True)
    if no_tiket is not None:
        ws.cell(row=baris_openpyxl, column=ct, value=no_tiket)
    if status is not None:
        ws.cell(row=baris_openpyxl, column=cs, value=status)
    ws.cell(row=baris_openpyxl, column=cw,
            value=datetime.now().isoformat(timespec="seconds"))
    try:
        wb.save(excel_path)
        return True, "ok"
    except PermissionError:
        return False, "file Excel sedang dibuka - TUTUP Excel lalu coba lagi."
