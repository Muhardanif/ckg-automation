"""
Tambah form khusus DEWASA ke PEMETAAN_PELAYANAN_FULL.xlsx (idempoten).

Form dewasa = sama dgn lansia MINUS SKILAS geriatri, PLUS 'Demografi Dewasa
Perempuan' yg punya pertanyaan tambahan '2. Apakah Anda sedang hamil?'. Form
anamnesis/klinis lain namanya sama → sudah tercakup config (otomatis dipakai
utk lansia & dewasa). Engine pelayanan.py melewati form yg kartunya tak ada.

PAKAI (TUTUP Excel PEMETAAN dulu):
  venv\\Scripts\\python.exe tools\\tambah_form_dewasa.py
"""
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

P = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                 "data", "output", "PEMETAAN_PELAYANAN_FULL.xlsx")
HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="2E7D32")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)

# nama_form, sumber, rows[(sq,tipe,judul,opsi,default,sel,cat)]
FORMS = [
    ("Demografi Dewasa Perempuan", "Excel/Default", [
        (100, "radio", "1. Status Perkawinan",
         "Belum Menikah  |  Menikah  |  Cerai Mati  |  Cerai Hidup",
         "(dari Excel/alat)", "sq_100i / PPM00000172", "Status Perkawinan dari registrasi"),
        (101, "radio", "2. Apabila belum menikah/cerai, ada rencana menikah dalam 1 tahun?",
         "Ya  |  Tidak", "Tidak", "sq_101i / PPM00000174",
         "KONDISIONAL: muncul bila Status=Belum Menikah/Cerai; default Tidak (CEK)"),
        (102, "radio", "3. Apakah Anda sedang hamil?", "Ya  |  Tidak",
         "Tidak", "sq_102i / PPM00000173", "CEK: ubah ke Ya bila peserta hamil"),
        (103, "radio", "4. Apakah Anda penyandang disabilitas?",
         "Non disabilitas  |  Penyandang disabilitas",
         "Non disabilitas", "sq_103i / PPM00000299", ""),
    ]),
    # FRM000006 (di-dump live 2026-06-16, peserta ACHMAD). Hanya 2 pertanyaan;
    # disabilitas di sq_102 (BUKAN sq_103 spt versi perempuan), tanpa Q hamil.
    ("Demografi Dewasa Laki-Laki", "Excel/Default", [
        (100, "radio", "1. Status Perkawinan",
         "Belum Menikah  |  Menikah  |  Cerai Mati  |  Cerai Hidup",
         "(dari Excel/alat)", "sq_100i / PPM00000172", "Status Perkawinan dari registrasi"),
        (101, "radio", "2. Apabila belum menikah/cerai, ada rencana menikah dalam 1 tahun?",
         "Ya  |  Tidak", "Tidak", "sq_101i / PPM00000174",
         "KONDISIONAL: muncul bila Status=Belum Menikah/Cerai; default Tidak (CEK)"),
        (102, "radio", "3. Apakah Anda penyandang disabilitas?",
         "Non disabilitas  |  Penyandang disabilitas",
         "Non disabilitas", "sq_102i / PPM00000299", ""),
    ]),
    # Kartu CATIN (di-dump live 2026-06-16, peserta ARUM). Hanya muncul utk catin;
    # peserta non-catin tak punya kartu ini → engine skip (tak-berlaku). 1 DROPDOWN.
    ("Riwayat Imunisasi Tetanus (Catin)", "Default", [
        (100, "dropdown",
         "1. Apakah anda pernah mendapatkan imunisasi tetanus minimal 2 kali?",
         "Pernah imunisasi tetanus minimal dua kali  |  Pernah imunisasi tetanus satu kali"
         "  |  Pernah imunisasi tetanus tetapi tidak ingat berapa kali  |  "
         "Tidak tahu atau tidak ingat",
         "Pernah imunisasi tetanus minimal dua kali", "sq_100i",
         "CEK: status imunisasi tetanus catin; baseline = minimal 2x (status T lengkap)"),
    ]),
]


def main():
    wb = openpyxl.load_workbook(P)
    ov = wb["Daftar Form"]
    tambah = 0
    for nama, sumber, rows in FORMS:
        if nama[:31] in wb.sheetnames:
            print(f"  sheet '{nama}' sudah ada, lewati.")
            continue
        ws = wb.create_sheet(nama[:31])
        cols = ["No", "sq", "Tipe", "Pertanyaan", "Opsi", "Nilai Default (ISI)",
                "id base / PPM", "Catatan"]
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = HDR
            cell.fill = FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for i, (sq, tipe, judul, opsi, default, sel, cat) in enumerate(rows, 1):
            ws.append([i, sq, tipe, judul, opsi, default, sel, cat])
        for k, w in enumerate([5, 7, 10, 38, 52, 24, 26, 30], 1):
            ws.column_dimensions[chr(64 + k)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = WRAP
                cell.border = THIN
        ov.append([ov.max_row, nama, "Ya", sumber, len(rows), ws.title])
        print(f"  + ditambah: {nama} ({len(rows)} pertanyaan)")
        tambah += 1
    try:
        wb.save(P)
    except PermissionError:
        raise SystemExit("GAGAL simpan: TUTUP Excel PEMETAAN_PELAYANAN_FULL.xlsx lalu ulangi.")
    print(f"[OK] {tambah} form ditambahkan ke {P}")


if __name__ == "__main__":
    main()
