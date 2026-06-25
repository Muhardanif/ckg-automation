"""
JALANKAN BATCH: daftarkan BANYAK baris dari Excel ke portal SATUSEHAT (mode CDP).

Fitur penting:
- RESUMABLE / anti-dobel: kolom hasil 'No. Tiket', 'Status Daftar', 'Waktu Daftar'
  ditulis balik ke Excel setelah tiap baris. Baris yang 'No. Tiket'-nya SUDAH
  terisi otomatis DILEWATI saat dijalankan ulang (tidak didaftarkan dua kali).
- Pra-cek tiap baris (field wajib + konsistensi NIK vs Tgl Lahir/JK). Baris yang
  bermasalah ditandai & dilewati (tidak membuang waktu ditolak portal), kecuali
  pakai --paksa.

PERSIAPAN (sama seperti trial):
  1. Jalankan Chrome dengan --remote-debugging-port=9222, login manual ke portal,
     buka menu CKG Umum > Cari/Daftarkan Individu.
  2. TUTUP file Excel di aplikasi Excel (skrip ini menulis balik ke file itu).

PAKAI:
  venv\\Scripts\\python.exe tools\\jalankan_batch.py --excel data/input/template_pendaftaran.xlsx
  # opsi: --mulai 1 --jumlah 50 --kelompok dewasa --paksa
"""
import argparse
import asyncio
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl                                            # noqa: E402

from app.automation.ckg_bot import CKGBot, LewatiPesertaError  # noqa: E402
from app.automation import selectors as S                 # noqa: E402
from app.schema import KelompokUsia                       # noqa: E402
from app.readers import (baca_excel, validasi, cek_konsistensi_nik,   # noqa: E402
                         koreksi_tgl_dari_nik, koreksi_jk_dari_nik)
from app.excel_hasil import (KOL_TIKET, KOL_STATUS, KOL_WAKTU,       # noqa: E402
                             STATUS_TERMINAL, warnai_baris)


def log(msg):
    print(f"[BATCH] {msg}", flush=True)


def _ensure_kolom(ws, header_row_1based, nama):
    """Cari kolom dgn header `nama` di baris header; buat baru bila belum ada.
    Kembalikan index kolom (1-based)."""
    last_col = ws.max_column
    for c in range(1, last_col + 1):
        v = ws.cell(row=header_row_1based, column=c).value
        if v is not None and str(v).strip() == nama:
            return c
    # buat kolom baru di ujung
    col = last_col + 1
    ws.cell(row=header_row_1based, column=col, value=nama)
    return col


def _simpan(wb, path):
    """Simpan workbook; beri pesan jelas bila file terkunci (masih dibuka Excel)."""
    try:
        wb.save(path)
        return True
    except PermissionError:
        log(f"GAGAL menyimpan '{path}': file sedang dibuka di Excel. "
            f"TUTUP Excel lalu jalankan lagi (baris yang sudah sukses tidak "
            f"akan diulang).")
        return False


async def _reset_ke_list(bot, page):
    """Pastikan kembali ke halaman list (tombol 'Daftar Baru' terlihat) sebelum
    memproses peserta berikutnya - menutup modal sisa dari peserta sebelumnya."""
    for _ in range(6):
        if await page.get_by_role("button", name="Daftar Baru").count() > 0:
            try:
                if await page.get_by_role(
                        "button", name="Daftar Baru").first.is_visible():
                    return True
            except Exception:
                pass
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(300)
    return await page.get_by_role("button", name="Daftar Baru").count() > 0


async def jalankan(args):
    # 1) baca data ternormalisasi (Peserta) + buka workbook utk tulis-balik
    ps = baca_excel(args.excel, KelompokUsia(args.kelompok),
                    header_row=args.header_row)
    if not ps:
        raise SystemExit(f"Tidak ada data terbaca di {args.excel}.")

    wb = openpyxl.load_workbook(args.excel)
    ws = wb.worksheets[0]
    header_row_1b = args.header_row + 1
    c_tiket = _ensure_kolom(ws, header_row_1b, KOL_TIKET)
    c_status = _ensure_kolom(ws, header_row_1b, KOL_STATUS)
    c_waktu = _ensure_kolom(ws, header_row_1b, KOL_WAKTU)
    # cek bisa tulis ke file (gagal cepat bila Excel masih terbuka)
    if not _simpan(wb, args.excel):
        return 2

    def _tandai(row, status, no_tiket=None):
        """Tulis Status + Waktu (dan No. Tiket bila ada) lalu warnai barisnya."""
        if no_tiket is not None:
            ws.cell(row=row, column=c_tiket, value=no_tiket)
        ws.cell(row=row, column=c_status, value=status)
        ws.cell(row=row, column=c_waktu,
                value=datetime.now().isoformat(timespec="seconds"))
        # warnai dari kolom pertama s/d kolom hasil terakhir (Waktu Daftar)
        warnai_baris(ws, row, status, kolom_akhir=c_waktu)

    # 2) tentukan rentang baris
    mulai_idx = max(args.mulai - 1, 0)
    akhir_idx = len(ps) if args.jumlah <= 0 else min(mulai_idx + args.jumlah, len(ps))
    target = ps[mulai_idx:akhir_idx]
    log(f"Total {len(ps)} baris; memproses {len(target)} "
        f"(baris data {mulai_idx + 1}..{akhir_idx}).")

    # 3) sambung ke Chrome (sekali)
    bot = CKGBot(headless=False, delay_ms=args.delay, cdp_url=args.cdp)
    try:
        await bot.connect_to_browser()
    except Exception as e:
        log(f"GAGAL connect ke Chrome: {type(e).__name__}: {e}")
        log("Pastikan Chrome jalan dgn --remote-debugging-port=9222 & sudah login.")
        return 1
    page = bot._page

    n_sukses = n_gagal = n_lewat = 0
    for p in target:
        row = p.baris_sumber  # = baris openpyxl (1-based)
        label = f"baris {row} | NIK={p.nik or '-'} | {p.nama or '-'}"

        # a) sudah pernah terdaftar? (No. Tiket terisi) -> lewati
        tiket_lama = ws.cell(row=row, column=c_tiket).value
        if tiket_lama is not None and str(tiket_lama).strip():
            log(f"LEWATI {label}: sudah ada No. Tiket {tiket_lama}.")
            n_lewat += 1
            continue

        # a2) status terminal (sudah CKG / data tidak valid) -> lewati, jangan
        #     proses ulang. (Untuk coba lagi: kosongkan sel Status baris itu.)
        status_lama = ws.cell(row=row, column=c_status).value
        if status_lama is not None and \
                str(status_lama).strip().startswith(STATUS_TERMINAL):
            log(f"LEWATI {label}: {str(status_lama).strip()}.")
            n_lewat += 1
            continue

        # b) pra-cek field wajib + konsistensi NIK
        err = validasi(p)
        warn = cek_konsistensi_nik(p)
        if err:
            pesan = "ERROR data: " + "; ".join(err)
            log(f"LEWATI {label}: {pesan}")
            _tandai(row, pesan)
            _simpan(wb, args.excel)
            n_lewat += 1
            continue

        # b2) Koreksi otomatis Tgl Lahir & Jenis Kelamin agar sesuai NIK (default
        #     aktif). NIK = kunci pencocokan Dukcapil, jadi data dari NIK lebih
        #     mungkin lolos. Daripada melewati baris, pakai data NIK lalu lanjut.
        dikoreksi = []
        if not args.no_koreksi_tgl:
            baru = koreksi_tgl_dari_nik(p)
            if baru and baru != p.tgl_lahir:
                log(f"KOREKSI {label}: Tgl Lahir {p.tgl_lahir or '-'} -> {baru} "
                    f"(sesuai NIK).")
                p.tgl_lahir = baru
                dikoreksi.append("tgl")
        if not args.no_koreksi_jk:
            jk = koreksi_jk_dari_nik(p)
            if jk and jk != p.jenis_kelamin:
                log(f"KOREKSI {label}: Jenis Kelamin {p.jenis_kelamin or '-'} -> "
                    f"{jk} (sesuai NIK).")
                p.jenis_kelamin = jk
                dikoreksi.append("jk")
        if dikoreksi:
            warn = cek_konsistensi_nik(p)       # nilai ulang; warning terkait hilang

        if warn and not args.paksa:
            pesan = "DILEWATI (cek data): " + "; ".join(warn)
            log(f"LEWATI {label}: {pesan} [pakai --paksa utk tetap coba]")
            _tandai(row, pesan)
            _simpan(wb, args.excel)
            n_lewat += 1
            continue
        if warn:
            log(f"PERINGATAN {label}: " + "; ".join(warn) + " (--paksa: tetap coba)")

        # c) daftarkan
        log(f"PROSES {label} ...")
        if not await _reset_ke_list(bot, page):
            log("Tidak bisa kembali ke halaman list. Hentikan batch.")
            break
        try:
            no_tiket = await bot.daftar_satu(
                p, on_step=lambda nama, info="": None)
            status = ("SUKSES (dikoreksi dari NIK: " + "+".join(dikoreksi) + ")"
                      if dikoreksi else "SUKSES")
            _tandai(row, status, no_tiket=no_tiket)
            log(f"SUKSES {label} -> No. Tiket {no_tiket}")
            n_sukses += 1
        except LewatiPesertaError as e:
            # Bukan gagal teknis (sudah pernah CKG / data tidak valid). Tandai
            # (terminal: tidak diproses ulang saat rerun), refresh halaman utk
            # menutup popup, lalu lanjut baris berikutnya. Tulis label ringkas
            # saja ('SUDAH CKG' / 'DATA TIDAK VALID'), tanpa detail.
            pesan = e.status_prefix
            _tandai(row, pesan)
            log(f"LEWATI {label}: {pesan}")
            n_lewat += 1
            try:
                await page.reload()
                await page.wait_for_load_state("networkidle")
            except Exception:
                pass
        except Exception as e:
            pesan = f"GAGAL: {str(e)[:300]}"
            _tandai(row, pesan)
            log(f"GAGAL {label}: {type(e).__name__}: {str(e)[:200]}")
            n_gagal += 1
        _simpan(wb, args.excel)

    await bot.stop()
    log("=" * 55)
    log(f"Selesai. Sukses={n_sukses}  Gagal={n_gagal}  Dilewati={n_lewat}")
    log(f"Hasil ditulis balik ke: {args.excel}")
    log("=" * 55)
    return 0


def main():
    ap = argparse.ArgumentParser(description="Batch pendaftaran CKG dari Excel (CDP).")
    ap.add_argument("--excel", required=True,
                    help="Path file Excel data (hasil ditulis balik ke file ini).")
    ap.add_argument("--kelompok", default="dewasa",
                    help="bayi/balita/dewasa/lansia (default dewasa).")
    ap.add_argument("--header-row", dest="header_row", type=int, default=0,
                    help="Indeks baris header (0 = baris pertama).")
    ap.add_argument("--mulai", type=int, default=1,
                    help="Mulai dari baris data ke-berapa (1-based). Default 1.")
    ap.add_argument("--jumlah", type=int, default=0,
                    help="Berapa baris diproses (0 = semua sampai akhir).")
    ap.add_argument("--delay", type=int, default=800,
                    help="Jeda antar-aksi (ms). Naikkan bila koneksi lambat.")
    ap.add_argument("--paksa", action="store_true",
                    help="Tetap coba walau pra-cek NIK memberi peringatan.")
    ap.add_argument("--no-koreksi-tgl", dest="no_koreksi_tgl",
                    action="store_true",
                    help="Matikan koreksi otomatis Tgl Lahir dari NIK (default: "
                         "aktif - tgl yg tak cocok NIK dibetulkan, bukan dilewati).")
    ap.add_argument("--no-koreksi-jk", dest="no_koreksi_jk",
                    action="store_true",
                    help="Matikan koreksi otomatis Jenis Kelamin dari NIK (default: "
                         "aktif - JK yg tak cocok NIK dibetulkan, bukan dilewati).")
    ap.add_argument("--cdp", default=S.CDP_URL,
                    help=f"URL Chrome remote-debugging (default {S.CDP_URL}).")
    args = ap.parse_args()
    sys.exit(asyncio.run(jalankan(args)))


if __name__ == "__main__":
    main()
