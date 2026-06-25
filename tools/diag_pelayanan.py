"""
DIAGNOSTIK halaman PELAYANAN (/ckg-pelayanan, mode CDP).

Tujuan: memetakan struktur halaman daftar pelayanan DAN form input hasil
pemeriksaan (di balik tombol 'Mulai'), agar selector + daftar field di
`app/automation/selectors.py` (PELAYANAN) dan `app/schema.py` (FIELD_PEMERIKSAAN)
bisa disesuaikan dengan portal asli (bukan placeholder).

Tanggal: secara default TIDAK disentuh (rentang bawaan UI biasanya sudah
mencakup tanggal hadir). Pakai --set-tanggal untuk memaksa set dari 'Waktu Hadir'
Excel (atau --tanggal YYYY-MM-DD). Filter pelayanan berupa RENTANG (mx-datepicker
2 panel); penyetelan best-effort dgn klik tanggal sama 2x lalu Escape.

Pencarian: dropdown filter di-set ke NIK (default), fallback ke Nama bila NIK
gagal. Lalu klik 'Mulai' (kolom Pemeriksaan) pada baris yang cocok, dan dump form.

Catatan: 'Mulai' hanya MEMBUKA form (status berubah saat data disimpan), jadi
dump ini tidak mengubah status peserta selama tidak menekan tombol simpan.

PERSIAPAN: Chrome jalan dgn --remote-debugging-port=9222 (1_mulai_chrome.bat),
sudah login portal SATUSEHAT.

PAKAI:
  venv\\Scripts\\python.exe tools\\diag_pelayanan.py --excel data\\input\\template_pendaftaran.xlsx --nik 3525084710630003
  # opsi: --pakai nama (cari via Nama), --set-tanggal, --no-mulai
"""
import argparse
import asyncio
import os
import re
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl                                             # noqa: E402
from app.automation import selectors as S                  # noqa: E402
from app.automation.ckg_bot import CKGBot                   # noqa: E402
from app.excel_hasil import KOL_WAKTU_HADIR                 # noqa: E402

URL_PELAYANAN = "https://sehatindonesiaku.kemkes.go.id/ckg-pelayanan"

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "data", "output")

# kandidat teks yang mungkin sedang terpilih di dropdown filter
KANDIDAT_FILTER = ("Nomor Tiket", "Nama Lengkap", "Nama", "Name", "NIK")

JS_DUMP_FORM = r"""() => {
  const visible = (el) => {
    const r = el.getBoundingClientRect();
    const st = getComputedStyle(el);
    return r.width > 0 && r.height > 0 && st.visibility !== 'hidden' && st.display !== 'none';
  };
  const txt = (el) => (el ? (el.textContent || '').replace(/\s+/g, ' ').trim() : '');
  const labelFor = (el) => {
    if (el.id) {
      const l = document.querySelector(`label[for="${CSS.escape(el.id)}"]`);
      if (l && txt(l)) return txt(l);
    }
    let p = el.closest('label');
    if (p && txt(p)) return txt(p).slice(0, 80);
    let node = el;
    for (let i = 0; i < 4 && node; i++) {
      let sib = node.previousElementSibling;
      while (sib) {
        const t = txt(sib);
        if (t && t.length < 80 && !sib.querySelector('input,select,textarea,button')) return t;
        sib = sib.previousElementSibling;
      }
      node = node.parentElement;
    }
    return '';
  };
  const inputs = Array.from(document.querySelectorAll('input')).filter(visible).map(i => ({
    label: labelFor(i), id: i.id, name: i.name, type: i.type,
    value: i.value, checked: i.checked, placeholder: i.placeholder
  }));
  const selects = Array.from(document.querySelectorAll('select')).filter(visible).map(s => ({
    label: labelFor(s), id: s.id, name: s.name,
    options: Array.from(s.options).map(o => o.text.trim()).slice(0, 40)
  }));
  const textareas = Array.from(document.querySelectorAll('textarea')).filter(visible).map(t => ({
    label: labelFor(t), id: t.id, name: t.name, placeholder: t.placeholder, value: t.value
  }));
  const customDD = Array.from(document.querySelectorAll("div[class*='cursor-pointer']"))
    .filter(visible).map(d => txt(d)).filter(t => t && t.length < 60);
  const buttons = Array.from(document.querySelectorAll("button,[role='button']"))
    .filter(visible).map(b => txt(b)).filter(Boolean);
  const headings = Array.from(document.querySelectorAll("h1,h2,h3,h4,h5,legend,[role='tab']"))
    .filter(visible).map(h => txt(h)).filter(t => t && t.length < 60);
  return {inputs, selects, textareas, customDD, buttons, headings, url: location.href};
}"""

# JS: judul tiap pertanyaan SurveyJS -> {sq_number: title}. Pasangkan elemen ber-id
# sq_NNN(i) ke kontainer 'sd-question' lalu ambil teks judulnya (bersih).
JS_QTITLE = r"""() => {
  const txt = (e) => (e.textContent || '').replace(/\s+/g, ' ').trim();
  const res = {};
  document.querySelectorAll("[id^='sq_']").forEach(el => {
    const m = el.id.match(/sq_(\d+)/); if (!m) return;
    const sq = m[1]; if (res[sq]) return;
    const q = el.closest("[class*='sd-question'],[class*='sv-question'],[data-name]");
    if (!q) return;
    const t = q.querySelector("[class*='question__title'],[class*='sd-title'],h5,h4,[class*='sv-title']");
    if (t) {
      let s = txt(t).replace(/Clear\s*More\s*No data to display$/i, '')
                    .replace(/No data to display$/i, '').trim();
      if (s) res[sq] = s;
    }
  });
  return res;
}"""

JS_DUMP_TABEL = r"""() => {
  const txt = (el) => (el ? (el.textContent || '').replace(/\s+/g, ' ').trim() : '');
  const tables = Array.from(document.querySelectorAll('table')).map(t => ({
    headers: Array.from(t.querySelectorAll('thead th, thead td')).map(txt),
    rows: Array.from(t.querySelectorAll('tbody tr')).slice(0, 5).map(tr =>
      Array.from(tr.querySelectorAll('td,th')).map(txt))
  }));
  return {tables};
}"""

JS_OPSI_DROPDOWN = r"""() => {
  const v = (e) => { const r = e.getBoundingClientRect(); return r.width > 0 && r.height > 0; };
  const sel = "div[class*='cursor-pointer'],[role='option'],li,div[class*='hover:bg']";
  return Array.from(new Set(Array.from(document.querySelectorAll(sel)).filter(v)
    .map(e => (e.textContent || '').replace(/\s+/g, ' ').trim())
    .filter(t => t && t.length < 40)));
}"""

# JS: daftar pemeriksaan di halaman detail (pasangkan tiap tombol aksi dgn judul
# kartu/baris terdekat) -> enumerasi 9 form + label tombolnya.
JS_DUMP_PEMERIKSAAN = r"""() => {
  const txt = (el) => (el ? (el.textContent || '').replace(/\s+/g, ' ').trim() : '');
  const vis = (el) => { const r = el.getBoundingClientRect(); return r.width > 0 && r.height > 0; };
  const btns = Array.from(document.querySelectorAll("button,a")).filter(vis)
    .filter(b => /input data|isi|lihat|edit|lanjut|mulai/i.test(txt(b)) && txt(b).length < 30);
  return btns.map(b => {
    let node = b, title = '';
    for (let i = 0; i < 6 && node; i++) {
      node = node.parentElement; if (!node) break;
      const t = txt(node);
      if (t && t.length > 12 && t.length < 160) { title = t; break; }
    }
    return { button: txt(b), title };
  });
}"""


def out(fh, msg=""):
    fh.write(msg + "\n")                 # file utf-8: selalu berhasil
    try:                                 # konsol Windows (cp1252) bisa gagal utk ≥, dll.
        print(msg, flush=True)
    except UnicodeEncodeError:
        print(msg.encode("ascii", "replace").decode("ascii"), flush=True)


def _tgl_iso(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return None


def _info_dari_excel(excel, nik, header_row=0):
    """Cari baris ber-NIK, kembalikan (iso_waktu_hadir, nama). (None, None) bila gagal."""
    try:
        wb = openpyxl.load_workbook(excel, data_only=True)
    except Exception as e:
        print(f"[DIAG] gagal buka Excel: {e}")
        return None, None
    ws = wb.worksheets[0]
    hdr = header_row + 1
    c_nik = c_hadir = c_nama = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        h = str(v).strip() if v is not None else ""
        if h.upper() == "NIK":
            c_nik = c
        elif h == KOL_WAKTU_HADIR:
            c_hadir = c
        elif h.lower() in ("nama", "nama lengkap"):
            c_nama = c
    if c_nik is None:
        wb.close()
        return None, None
    for r in range(hdr + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=c_nik).value
        if v is not None and str(v).strip() == str(nik):
            iso = _tgl_iso(ws.cell(row=r, column=c_hadir).value) if c_hadir else None
            nama = ws.cell(row=r, column=c_nama).value if c_nama else None
            wb.close()
            return iso, (str(nama).strip() if nama else None)
    wb.close()
    return None, None


async def _tutup_popup(page):
    """Tutup popup kalender / overlay agar tak menghalangi klik berikutnya."""
    try:
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(250)
    except Exception:
        pass


async def _set_tanggal_rentang(bot, fh, iso_date):
    from datetime import datetime as _dt
    try:
        d = _dt.strptime(iso_date, "%Y-%m-%d")
        root = bot._page.locator(".mx-datepicker").first
        if await root.count() == 0:
            out(fh, "[DIAG] .mx-datepicker tak ada.")
            return
        await bot._pilih_tanggal_mx(root, d)
        await bot._page.wait_for_timeout(400)
        await bot._pilih_tanggal_mx(root, d)
        out(fh, f"[DIAG] Tanggal di-set (rentang {iso_date}..{iso_date}).")
    except Exception as e:
        out(fh, f"[DIAG] set tanggal gagal: {e}")
    await _tutup_popup(bot._page)


# Selektor opsi di panel dropdown Vue portal CKG (mengambang). Disalin dari pola
# terbukti app/automation/ckg_bot.py::_pilih_dropdown: opsi = div cursor-pointer
# +hover:bg-gray-1 di panel shadow-standard. Fallback [role=option]/li.
_DD_OPT_SEL = [
    "div[class*='cursor-pointer'][class*='hover:bg-gray-1']",
    "div[class*='shadow-standard'] div[class*='cursor-pointer']",
    "[role='option']", "li[role='option']",
]


async def _kontrol_filter(page, label):
    """Locator kontrol dropdown filter ber-teks `label` (mis. 'Nama').

    Targetkan div ber-class 'cursor-pointer' yg teksnya PERSIS `label` — ini
    kotak dropdown filter sebenarnya. Hindari `get_by_text` global yg bisa kena
    HEADER KOLOM tabel ('Nama'/'Nomor Tiket') → klik di situ tak membuka popup."""
    pat = re.compile(rf"^\s*{re.escape(label)}\s*$")
    cand = page.locator("div[class*='cursor-pointer']").filter(has_text=pat).locator("visible=true")
    if await cand.count() > 0:
        return cand.first
    # fallback: teks label HARUS punya leluhur 'cursor-pointer' (itu kontrol klik).
    # Tanpa leluhur cursor-pointer = HEADER KOLOM tabel ('Nomor Tiket'/'Nama') atau
    # teks biasa → BUKAN kontrol dropdown → tolak (return None, coba kandidat lain).
    loc = page.get_by_text(pat).locator("visible=true").first
    if await loc.count() > 0:
        anc = loc.locator("xpath=ancestor-or-self::*[contains(@class,'cursor-pointer')][1]")
        if await anc.count() > 0:
            return anc.first
    return None


async def _set_dropdown_filter(page, fh, target):
    """Setel dropdown filter ke `target` (mis. 'NIK'/'Nama'). Return teks final
    atau None. Buka panel, verifikasi opsi muncul, klik opsi `target`."""
    await _tutup_popup(page)
    # 1) tentukan label filter saat ini (kandidat pertama yg ADA sbg kontrol).
    ctrl, cur = None, None
    for c in KANDIDAT_FILTER:
        k = await _kontrol_filter(page, c)
        if k is not None and await k.count() > 0:
            ctrl, cur = k, c
            break
    if ctrl is None:
        out(fh, "[DIAG] Kontrol dropdown filter tak ditemukan.")
        return None
    if cur == target:
        out(fh, f"[DIAG] Dropdown filter sudah '{target}'.")
        return target
    out(fh, f"[DIAG] Dropdown filter saat ini '{cur}' -> ganti ke '{target}'.")
    # 2) buka panel + klik opsi target. Retry buka bila panel belum render.
    opt_exact = re.compile(rf"^\s*{re.escape(target)}\s*$", re.I)
    for attempt in range(3):
        try:
            await ctrl.click(timeout=5000)
        except Exception as e:
            out(fh, f"[DIAG] gagal klik kontrol (terhalang?): {e}")
            await _tutup_popup(page)
            continue
        await page.wait_for_timeout(500)
        for psel in _DD_OPT_SEL:
            opt = page.locator(psel).filter(has_text=opt_exact).locator("visible=true")
            try:
                if await opt.count() > 0:
                    await opt.first.click(timeout=4000)
                    await page.wait_for_timeout(400)
                    out(fh, f"[DIAG] Dropdown filter -> {target}.")
                    return target
            except Exception:
                continue
        await page.wait_for_timeout(300)
    # 3) gagal: dump opsi terlihat + HTML kontrol utk diagnosis.
    opsi = await page.evaluate(JS_OPSI_DROPDOWN)
    out(fh, f"[DIAG] Opsi dropdown terlihat: {opsi}")
    try:
        html = await ctrl.evaluate(
            "el => { const w = el.closest('div')||el; return (w.outerHTML||'').slice(0,800); }")
        out(fh, "[DIAG] HTML kontrol filter: " + " ".join(html.split()))
    except Exception:
        pass
    out(fh, f"[DIAG] Opsi '{target}' tak ada setelah dropdown dibuka.")
    return None


async def _cari(page, fh, kata, match_text=None, tries=16):
    """Ketik `kata` ke kotak cari lalu poll sampai baris hasil muncul.

    `match_text` (opsional): teks yang dipoll di baris hasil. Berguna saat kata
    kunci pencarian TAK tampil di baris (mis. filter NIK → baris cuma tampilkan
    Nama): cari pakai NIK tapi poll pakai Nama. Default: poll `kata` itu sendiri.
    `tries`: jumlah polling x500ms. Pakai nilai KECIL utk pencarian awal sebelum
    pemilihan tab (peserta bisa di tab lain → tunggu lama di tab salah = sia-sia)."""
    cek = str(match_text) if match_text is not None else str(kata)
    box = page.get_by_placeholder(re.compile(r"Masukkan|cari|nik|nama", re.I)).first
    if await box.count() == 0:
        box = page.get_by_role("textbox").last
    if await box.count() == 0:
        out(fh, "[DIAG] Kotak cari tak ditemukan.")
        return False
    await box.click(timeout=6000)
    try:
        await box.fill("", timeout=3000)
    except Exception:
        pass
    # KETIK per-karakter (bukan fill): kotak 'searchNik' (filter NIK) MENOLAK value
    # yg di-set sekaligus → error 'NIK hanya bisa angka'. Mengetik tiap karakter
    # memicu handler input Vue dgn benar (terbukti via tools/diag_nik_box.py).
    await box.type(str(kata), delay=80)
    await box.press("Enter")
    # poll: tunggu tabel benar2 terfilter (baris memuat `cek`) hingga tries x500ms.
    target = page.locator("tbody tr").filter(has_text=cek)
    for _ in range(max(1, tries)):
        await page.wait_for_timeout(500)
        try:
            if await target.count() > 0:
                out(fh, f"[DIAG] (cari) hasil '{kata}' muncul (cocok '{cek}').")
                return True
        except Exception:
            pass
    out(fh, f"[DIAG] (cari) '{kata}' belum muncul setelah ~{max(1, tries)*0.5:.0f} detik "
            f"(peserta mungkin di tab lain / filter lambat).")
    return True


# Sentinel: baris peserta menandakan transaksi pelayanan SUDAH SELESAI
# (Pemeriksaan Mandiri 'Lengkap' + Pelayanan 'Selesai Pemeriksaan'). Dipakai
# pemanggil utk membedakan "selesai, jangan diklik" dari "tak ketemu" (None).
SUDAH_SELESAI = object()


async def _baris_transaksi_selesai(page, term):
    """True bila baris listing utk `term` menandakan transaksi SUDAH SELESAI:
    kolom 'Pemeriksaan Mandiri' = 'Lengkap' (BUKAN 'Belum Lengkap') DAN kolom
    'Pelayanan' = 'Selesai Pemeriksaan' (dua label hijau). HANYA membaca isi
    baris, tidak mengklik apa pun."""
    rows = page.locator("tbody tr").filter(has_text=str(term))
    if await rows.count() == 0:
        return False
    try:
        teks = (await rows.first.inner_text()).lower()
    except Exception:
        return False
    mandiri_lengkap = ("lengkap" in teks) and ("belum lengkap" not in teks)
    pelayanan_selesai = "selesai pemeriksaan" in teks
    return mandiri_lengkap and pelayanan_selesai


async def _klik_mulai(page, fh, term, aksi="Mulai"):
    """Buka detail dari baris yg mengandung `term`. Cetak tombol di baris lalu
    coba kandidat label (aksi -> Lanjutkan -> Mulai). Hindari 'Lihat' (=Rapor)."""
    rows = page.locator("tbody tr").filter(has_text=str(term))
    n = await rows.count()
    out(fh, f"[DIAG] Baris cocok '{term}': {n}.")
    if n == 0:
        # JANGAN klik baris lain (bisa buka peserta salah). Tampilkan baris yang ada.
        nama_baris = []
        allrows = page.locator("tbody tr")
        for i in range(min(await allrows.count(), 8)):
            try:
                c = allrows.nth(i).locator("td").nth(1)
                nama_baris.append((await c.inner_text()).strip())
            except Exception:
                pass
        out(fh, f"[DIAG] Pencarian '{term}' tak menghasilkan baris. Baris terlihat: {nama_baris}. "
                f"Filter mungkin belum jalan / peserta di tab lain. BERHENTI (tak klik apa pun).")
        return False
    row = rows.first
    rb = row.get_by_role("button")
    labels = []
    for i in range(min(await rb.count(), 10)):
        try:
            labels.append((await rb.nth(i).inner_text()).strip())
        except Exception:
            pass
    out(fh, f"[DIAG] Tombol di baris: {labels}")
    for k in (aksi, "Lanjutkan", "Mulai"):
        pat = re.compile(rf"^\s*{re.escape(k)}\s*$", re.I)
        btn = row.get_by_role("button", name=pat)
        if await btn.count() == 0:
            btn = row.get_by_text(pat).locator("visible=true")
        if await btn.count() > 0:
            await btn.first.click(timeout=8000)
            await page.wait_for_timeout(3000)
            try:
                await page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass
            out(fh, f"[DIAG] Klik aksi baris '{k}'.")
            return True
    out(fh, "[DIAG] Tak ada tombol aksi cocok (Mulai/Lanjutkan) di baris.")
    return False


def _dump_form(form, fh, judul):
    """Cetak hasil page.evaluate(JS_DUMP_FORM) dengan judul."""
    out(fh, "=" * 70)
    out(fh, f"[{judul}] URL: {form.get('url')}")
    out(fh, f"[{judul}] INPUT (label | id | name | type | value | placeholder):")
    for i in form["inputs"]:
        out(fh, f"   - {i['label']!r:40} | id={i['id']!r} name={i['name']!r} "
                f"type={i['type']!r} checked={i['checked']} "
                f"value={i['value']!r} ph={i['placeholder']!r}")
    if form.get("qtitle"):
        out(fh, f"[{judul}] QTITLE (sq -> judul pertanyaan):")
        for sq, t in sorted(form["qtitle"].items(), key=lambda kv: int(kv[0])):
            out(fh, f"   - sq_{sq}: {t!r}")
    out(fh, f"[{judul}] SELECT (label | id | name | options):")
    for s in form["selects"]:
        out(fh, f"   - {s['label']!r} id={s['id']!r} name={s['name']!r}")
        out(fh, f"       options: {s['options']}")
    out(fh, f"[{judul}] TEXTAREA:")
    for t in form["textareas"]:
        out(fh, f"   - {t['label']!r} id={t['id']!r} name={t['name']!r} ph={t['placeholder']!r}")
    out(fh, f"[{judul}] Dropdown custom (Vue):")
    for t in dict.fromkeys(form["customDD"]):
        out(fh, f"   - {t!r}")
    out(fh, f"[{judul}] Heading/tab/section:")
    for t in dict.fromkeys(form["headings"]):
        out(fh, f"   - {t!r}")
    out(fh, f"[{judul}] Tombol:")
    for t in dict.fromkeys(form["buttons"]):
        out(fh, f"   - {t!r}")


async def _pilih_tab(page, fh, nama):
    """Klik tab status listing (Belum/Sedang/Selesai Pemeriksaan)."""
    if not nama:
        return
    await _tutup_popup(page)
    loc = page.get_by_text(re.compile(rf"^\s*{re.escape(nama)}(\s*\d+)?\s*$"), exact=False).locator("visible=true")
    if await loc.count() == 0:
        out(fh, f"[DIAG] Tab '{nama}' tak ditemukan.")
        return
    try:
        await loc.first.click(timeout=5000)
        await page.wait_for_timeout(1500)
        out(fh, f"[DIAG] Pindah tab -> {nama}.")
    except Exception as e:
        out(fh, f"[DIAG] gagal klik tab '{nama}': {e}")


def _dump_pemeriksaan(items, fh):
    out(fh, "[PEMERIKSAAN] Daftar form (tombol -> judul kartu terdekat):")
    seen = set()
    for it in items:
        key = (it["button"], it["title"])
        if key in seen:
            continue
        seen.add(key)
        out(fh, f"   - [{it['button']}] {it['title']!r}")


async def _tunggu_overlay_hilang(page, timeout_ms=8000):
    """Tunggu overlay modal (div fixed inset-0 z-1000) menghilang."""
    ov = page.locator("div[class*='fixed'][class*='inset-0']")
    waited = 0
    while waited < timeout_ms:
        try:
            vis = (await ov.count() > 0) and await ov.first.is_visible()
        except Exception:
            vis = False
        if not vis:
            return True
        await page.wait_for_timeout(300)
        waited += 300
    return False


async def _lewati_modal_mulai(page, fh, iso_periksa):
    """Tangani modal kalender setelah 'Mulai Pemeriksaan': pilih tanggal periksa
    (default hari ini) lalu klik 'Simpan'. Day-button hanya ada di modal ini."""
    from datetime import date as _date
    await page.wait_for_timeout(800)
    simpan = page.get_by_role("button", name=re.compile(r"^\s*Simpan\s*$", re.I))
    if await simpan.count() == 0:
        out(fh, "[DIAG] Modal 'Simpan' tak terdeteksi (mungkin tak ada modal). Lanjut.")
        return False
    try:
        d = datetime.strptime(iso_periksa, "%Y-%m-%d").day if iso_periksa else _date.today().day
    except (ValueError, TypeError):
        d = _date.today().day
    hari = page.get_by_role("button", name=re.compile(rf"^\s*{d}\s*$"))
    try:
        if await hari.count() > 0:
            await hari.first.click(timeout=4000)
            out(fh, f"[DIAG] Pilih tanggal periksa di modal: hari {d}.")
    except Exception as e:
        out(fh, f"[DIAG] gagal pilih hari {d} (mungkin sudah terpilih): {e}")
    try:
        await simpan.first.click(timeout=6000)
        out(fh, "[DIAG] Klik 'Simpan' modal mulai pemeriksaan.")
    except Exception as e:
        out(fh, f"[DIAG] gagal klik 'Simpan' modal: {e}")
        return False
    await _tunggu_overlay_hilang(page)
    await page.wait_for_timeout(1500)
    return True


async def _klik_teks(page, fh, teks, exact=False):
    """Klik tombol/elemen dgn teks `teks`. Return True bila berhasil."""
    pat = re.compile(rf"^\s*{re.escape(teks)}\s*$" if exact else re.escape(teks), re.I)
    cand = page.get_by_role("button", name=pat)
    if await cand.count() == 0:
        cand = page.get_by_text(pat).locator("visible=true")
    if await cand.count() == 0:
        out(fh, f"[DIAG] Elemen '{teks}' tak ditemukan untuk diklik.")
        return False
    try:
        await cand.first.click(timeout=8000)
        await page.wait_for_timeout(2500)
        try:
            await page.wait_for_load_state("networkidle", timeout=8000)
        except Exception:
            pass
        out(fh, f"[DIAG] Klik '{teks}' OK.")
        return True
    except Exception as e:
        out(fh, f"[DIAG] gagal klik '{teks}': {e}")
        return False


def _frag(judul):
    """Potongan judul yang aman utk pencocokan teks (buang nomor awal & kurung)."""
    s = re.sub(r"^\s*\d+[a-z]?\.\s*", "", judul)   # buang '1. ' / '6a. '
    s = s.split("(")[0].split("=>")[0].split(">=")[0].strip()
    return s[:40]


def _baca_ya_forms(path):
    """Baca sheet 'Daftar Form' di PEMETAAN; kembalikan list (nama, sumber) yg Otomasi=Ya."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Daftar Form"]
    hasil = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        nama, otomasi, sumber = r[3], r[4], r[5]
        if otomasi and str(otomasi).strip().lower().startswith("ya"):
            hasil.append((str(nama).strip(), str(sumber or "").strip()))
    wb.close()
    return hasil


async def _buka_form(page, fh, judul):
    """Buka satu form: klik tombol 'Input Data' pada kartu yg memuat `judul`.

    PENTING: kartu-kartu form dibungkus dalam container grup (mis. 9 anamnesis
    dalam satu kotak). Memakai 'leluhur pertama yg memuat judul' SALAH — leluhur
    grup memuat SEMUA judul, jadi selalu mengklik tombol kartu PERTAMA di grup
    (bug duplikat). Solusi: untuk tiap tombol cari leluhur TERKECIL yang memuat
    judul, lalu pilih tombol dgn leluhur tersempit (= kartu tunggal, bukan grup)."""
    frag = _frag(judul).lower()
    btns = page.get_by_role("button", name=re.compile(r"Input Data", re.I))
    # tunggu kartu 'Input Data' render (detail kadang lambat usai kembali dari form)
    nb = 0
    for _ in range(20):                     # ~10 detik
        nb = await btns.count()
        if nb > 0:
            break
        await page.wait_for_timeout(500)
    best_i, best_len = -1, 10 ** 9
    for i in range(nb):
        b = btns.nth(i)
        for lvl in range(2, 9):
            try:
                anc = b.locator(f"xpath=ancestor::*[{lvl}]")
                t = (await anc.inner_text()).replace("\n", " ").strip()
            except Exception:
                break
            if frag and frag in t.lower():
                if len(t) < best_len:   # leluhur terkecil utk tombol ini
                    best_len, best_i = len(t), i
                break                   # naik lebih tinggi hanya memperbesar
    if best_i < 0 or best_len > 200:    # >200 = leluhur grup, bukan kartu tunggal
        out(fh, f"[DIAG] Tombol 'Input Data' utk {judul!r} (frag={frag!r}) "
                f"tak ditemukan sbg kartu tunggal (min_len={best_len}).")
        return False
    try:
        await btns.nth(best_i).click(timeout=8000)
        await page.wait_for_timeout(2500)
        try:
            await page.wait_for_load_state("networkidle", timeout=8000)
        except Exception:
            pass
        return True
    except Exception as e:
        out(fh, f"[DIAG] gagal klik Input Data utk {judul!r}: {e}")
        return False


# JS: setelah dropdown diklik, kumpulkan kandidat opsi yg terlihat + outerHTML
# kontainer listbox terdekat (untuk belajar struktur widget vue-select).
JS_PROBE_OPSI = r"""() => {
  const vis = (e) => { const r = e.getBoundingClientRect(); const s = getComputedStyle(e);
    return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none'; };
  const txt = (e) => (e.textContent || '').replace(/\s+/g, ' ').trim();
  const sel = "[role='option'],li,[class*='option'],[class*='vs__dropdown'],"
            + "[class*='multiselect__'],[class*='dropdown-menu'],[class*='el-select'],"
            + "[class*='ant-select-item'],[class*='p-dropdown-item'],[class*='hover:bg']";
  const opts = Array.from(new Set(Array.from(document.querySelectorAll(sel))
    .filter(vis).map(txt).filter(t => t && t.length < 80)));
  // listbox container terdekat (yg punya banyak anak opsi)
  let box = null, bestN = 1;
  for (const c of document.querySelectorAll("ul,[role='listbox'],[class*='dropdown'],[class*='menu']")) {
    if (!vis(c)) continue;
    const n = c.querySelectorAll("li,[role='option'],[class*='option']").length;
    if (n > bestN) { bestN = n; box = c; }
  }
  return { opts, boxHTML: box ? box.outerHTML.slice(0, 1500) : '' };
}"""


async def _dump_opsi_dropdown(page, fh, judul):
    """PROBE: di halaman form, klik tiap input dropdown (text id^=sq_) & dump opsi
    yg muncul + outerHTML listbox (sekali jalan utk belajar struktur)."""
    out(fh, "-" * 70)
    out(fh, f"[OPSI] Form {judul!r} — probe dropdown:")
    inputs = page.locator("input[type='text'][id^='sq_']")
    n = await inputs.count()
    out(fh, f"[OPSI] dropdown text-input ditemukan: {n}")
    for i in range(n):
        inp = inputs.nth(i)
        try:
            iid = await inp.get_attribute("id")
            ph = await inp.get_attribute("placeholder")
        except Exception:
            iid, ph = "?", "?"
        # input vue-select sering ukuran 0 / tertutup → klik wrapper toggle (ancestor div).
        dibuka = False
        try:
            await inp.scroll_into_view_if_needed(timeout=3000)
        except Exception:
            pass
        for xp in ("xpath=ancestor::div[1]", "xpath=ancestor::div[2]", "xpath=ancestor::div[3]"):
            try:
                await inp.locator(xp).first.click(timeout=3000)
                await page.wait_for_timeout(700)
                probe = await page.evaluate(JS_PROBE_OPSI)
                if probe["opts"] or probe["boxHTML"]:
                    dibuka = True
                    break
            except Exception:
                continue
        if not dibuka:
            try:                                  # upaya terakhir: force-click input
                await inp.click(timeout=2000, force=True)
                await page.wait_for_timeout(700)
            except Exception as e:
                out(fh, f"[OPSI] id={iid!r} gagal buka dropdown: {e}")
                continue
        res = await page.evaluate(JS_PROBE_OPSI)
        out(fh, f"[OPSI] id={iid!r} ph={ph!r}")
        out(fh, f"       opsi: {res['opts']}")
        if res["boxHTML"]:
            out(fh, f"       boxHTML: {res['boxHTML']}")
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(300)


async def _kembali_ke_detail(page, fh):
    """Kembali dari form.kemkes.go.id ke halaman detail-pemeriksaan."""
    btn = page.get_by_role("button", name=re.compile(r"Kembali ke Halaman Utama", re.I))
    if await btn.count() == 0:
        btn = page.get_by_text(re.compile(r"Kembali ke Halaman Utama", re.I)).locator("visible=true")
    try:
        if await btn.count() > 0:
            await btn.first.click(timeout=8000)
        else:
            await page.go_back()
    except Exception:
        try:
            await page.go_back()
        except Exception:
            pass
    # tunggu sampai kembali ke detail-pemeriksaan
    for _ in range(20):
        await page.wait_for_timeout(400)
        if "detail-pemeriksaan" in (page.url or ""):
            await page.wait_for_timeout(800)
            return True
    out(fh, "[DIAG] PERINGATAN: belum kembali ke detail-pemeriksaan.")
    return False


async def main():
    ap = argparse.ArgumentParser(description="Diagnostik halaman pelayanan CKG.")
    ap.add_argument("--excel", default="", help="Excel hasil (ambil nama/'Waktu Hadir' per NIK).")
    ap.add_argument("--nik", default="", help="NIK contoh utk dibuka form-nya.")
    ap.add_argument("--pakai", choices=["nik", "nama"], default="nik",
                    help="Cari via NIK (default) atau Nama.")
    ap.add_argument("--tanggal", default="", help="Override tanggal YYYY-MM-DD (perlu --set-tanggal).")
    ap.add_argument("--set-tanggal", dest="set_tanggal", action="store_true",
                    help="Paksa set rentang tanggal (default: pakai rentang bawaan UI).")
    ap.add_argument("--header-row", dest="header_row", type=int, default=0)
    ap.add_argument("--no-mulai", dest="no_mulai", action="store_true")
    ap.add_argument("--mulai-pemeriksaan", dest="mulai_pemeriksaan", action="store_true",
                    help="Setelah buka detail, klik 'Mulai Pemeriksaan' + 'Input Data' lalu dump.")
    ap.add_argument("--seksi", default="",
                    help="Nama (sebagian) modul skrining Nakes utk dibuka & di-dump, "
                         "mis. \"Skrining Gizi\". Perlu --mulai-pemeriksaan.")
    ap.add_argument("--tanggal-periksa", dest="tanggal_periksa", default="",
                    help="Tanggal di modal 'Mulai Pemeriksaan' (YYYY-MM-DD). Default: hari ini.")
    ap.add_argument("--tab", default="",
                    help="Pindah tab listing dulu: 'Sedang Pemeriksaan' / 'Selesai Pemeriksaan'.")
    ap.add_argument("--aksi", default="Mulai",
                    help="Label tombol aksi baris utk buka detail (Mulai/Lanjutkan/Lihat).")
    ap.add_argument("--buka", default="",
                    help="Setelah detail terbuka, buka 1 form pemeriksaan dgn judul/teks ini & dump.")
    ap.add_argument("--buka-semua", dest="buka_semua", action="store_true",
                    help="Dump SEMUA form yg Otomasi=Ya di file --pemetaan (buka->dump->kembali).")
    ap.add_argument("--pemetaan", default=os.path.join(OUT_DIR, "PEMETAAN_PELAYANAN.xlsx"),
                    help="Path PEMETAAN_PELAYANAN.xlsx (sumber daftar form Ya).")
    ap.add_argument("--dump-opsi", dest="dump_opsi", action="store_true",
                    help="Setelah buka form (--buka/--buka-semua), klik tiap dropdown "
                         "& dump opsinya (probe widget vue-select).")
    args = ap.parse_args()

    os.makedirs(OUT_DIR, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(OUT_DIR, f"diag_pelayanan_{stamp}.txt")
    fh = open(out_path, "w", encoding="utf-8")

    iso, nama = (args.tanggal or None), None
    if args.excel and args.nik:
        iso_x, nama = _info_dari_excel(args.excel, args.nik, args.header_row)
        iso = iso or iso_x
        out(fh, f"[DIAG] Dari Excel: nama={nama!r}, Waktu Hadir={iso_x!r}")

    bot = CKGBot(headless=False, cdp_url=S.CDP_URL)
    try:
        await bot.connect_to_browser()
        page = bot._page
        # Listing = persis /ckg-pelayanan. Bila tab tertinggal di detail-pemeriksaan
        # (run sebelumnya), URL tetap memuat '/ckg-pelayanan' → paksa kembali ke listing.
        cur = page.url or ""
        if not cur.rstrip("/").endswith("/ckg-pelayanan"):
            out(fh, f"[DIAG] Navigasi ke {URL_PELAYANAN} (dari {cur})")
            await page.goto(URL_PELAYANAN)
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
        await page.wait_for_timeout(1500)
        out(fh, f"[DIAG] Tab aktif: {page.url}")
        out(fh, "=" * 70)

        if args.set_tanggal and iso:
            await _set_tanggal_rentang(bot, fh, iso)
        else:
            out(fh, "[DIAG] Tanggal tidak disentuh (pakai rentang bawaan UI).")
            await _tutup_popup(page)

        # pindah tab status bila diminta (mis. peserta sudah 'Sedang Pemeriksaan')
        await _pilih_tab(page, fh, args.tab)

        # 1) dump listing
        listing = await page.evaluate(JS_DUMP_FORM)
        out(fh, "[LISTING] Dropdown custom / cursor-pointer:")
        for t in dict.fromkeys(listing["customDD"]):
            out(fh, f"   - {t!r}")
        out(fh, "[LISTING] Tombol terlihat:")
        for t in dict.fromkeys(listing["buttons"]):
            out(fh, f"   - {t!r}")
        tab = await page.evaluate(JS_DUMP_TABEL)
        for ti, t in enumerate(tab["tables"]):
            out(fh, f"[LISTING] Tabel #{ti}: headers = {t['headers']}")
            for r in t["rows"][:3]:
                out(fh, f"      row: {r}")
        out(fh, "=" * 70)

        # 2) buka form
        if args.nik and not args.no_mulai:
            target = "NIK" if args.pakai == "nik" else "Nama"
            term = args.nik if args.pakai == "nik" else (nama or args.nik)
            res = await _set_dropdown_filter(page, fh, target)
            if res is None and args.pakai == "nik" and nama:
                out(fh, "[DIAG] NIK gagal -> fallback ke Nama.")
                res = await _set_dropdown_filter(page, fh, "Nama")
                term = nama
            await _cari(page, fh, term)
            if await _klik_mulai(page, fh, term, aksi=args.aksi):
                form = await page.evaluate(JS_DUMP_FORM)
                _dump_form(form, fh, "DETAIL")
                detail_url = page.url        # stabil per peserta; dipakai antar-form

                # bila peserta belum mulai: lakukan 'Mulai Pemeriksaan' + modal tanggal
                if args.mulai_pemeriksaan:
                    await _klik_teks(page, fh, "Mulai Pemeriksaan")
                    # tanggal modal: --tanggal-periksa > tanggal hadir (Excel) > hari ini
                    await _lewati_modal_mulai(page, fh, args.tanggal_periksa or iso)
                    await _tunggu_overlay_hilang(page)

                # enumerasi daftar 9 form pemeriksaan (tombol -> judul kartu)
                items = await page.evaluate(JS_DUMP_PEMERIKSAAN)
                _dump_pemeriksaan(items, fh)

                # buka 1 form pemeriksaan tertentu & dump (form ada di form.kemkes.go.id)
                if args.buka:
                    out(fh, f"[DIAG] Membuka form pemeriksaan: {args.buka!r}")
                    if await _buka_form(page, fh, args.buka):
                        formx = await page.evaluate(JS_DUMP_FORM)
                        formx["qtitle"] = await page.evaluate(JS_QTITLE)
                        _dump_form(formx, fh, f"FORM:{args.buka}")
                        if args.dump_opsi:
                            await _dump_opsi_dropdown(page, fh, args.buka)
                        await _kembali_ke_detail(page, fh)

                # dump SEMUA form Otomasi=Ya dari file pemetaan
                if args.buka_semua:
                    try:
                        ya = _baca_ya_forms(args.pemetaan)
                    except Exception as e:
                        out(fh, f"[DIAG] gagal baca pemetaan {args.pemetaan}: {e}")
                        ya = []
                    out(fh, f"[DIAG] Akan dump {len(ya)} form (Otomasi=Ya).")
                    for idx, (nama_form, sumber) in enumerate(ya, 1):
                        out(fh, "#" * 70)
                        out(fh, f"[DIAG] ({idx}/{len(ya)}) Buka form: {nama_form!r} [sumber={sumber}]")
                        # TAHAN BANTING: satu form gagal tak boleh menggugurkan sisanya.
                        try:
                            if await _buka_form(page, fh, nama_form):
                                formx = await page.evaluate(JS_DUMP_FORM)
                                try:
                                    formx["qtitle"] = await page.evaluate(JS_QTITLE)
                                except Exception as e:
                                    out(fh, f"[DIAG] qtitle gagal {nama_form!r}: {e}")
                                _dump_form(formx, fh, f"FORM:{nama_form}")
                                if args.dump_opsi:
                                    await _dump_opsi_dropdown(page, fh, nama_form)
                            else:
                                out(fh, f"[DIAG] LEWATI {nama_form!r} (tombol tak ketemu).")
                        except Exception as e:
                            out(fh, f"[DIAG] ERROR form {nama_form!r}: {e}")
                        # ANDAL: navigasi langsung ke URL detail (bukan tombol 'Kembali'
                        # yg sering balik ke detail kosong) lalu tunggu kartu render.
                        if idx < len(ya):
                            try:
                                await page.goto(detail_url)
                                await page.wait_for_load_state("networkidle", timeout=10000)
                            except Exception:
                                pass
                            ib = page.get_by_role("button", name=re.compile(r"Input Data", re.I))
                            for _ in range(20):
                                if await ib.count() > 0:
                                    break
                                await page.wait_for_timeout(500)

        out(fh, "=" * 70)
        out(fh, f"[DIAG] Selesai. Dump: {out_path}")
    finally:
        fh.close()
        try:
            await bot.stop()
        except Exception:
            pass


if __name__ == "__main__":
    asyncio.run(main())
