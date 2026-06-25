"""
Generator PEMETAAN_PELAYANAN_FULL.xlsx — workbook kerja LENGKAP untuk Step 3.

Membaca dump diagnostik (bukan tebakan) lalu menulis 1 sheet per form berisi
TIAP PERTANYAAN: sq id, tipe (radio/number/dropdown), opsi, kolom 'Nilai Default'
(KOSONG untuk diisi user), dan selector (id/PPM) untuk membangun otomasi nanti.

Sumber dump (peserta lansia perempuan SUMIATI, tab 'Sedang Pemeriksaan'):
  - FIELD bersih 22 form : data/output/diag_pelayanan_20260615_185327.txt
  - OPSI dropdown        : diag_pelayanan_20260615_{190348,203231,203257,203326,203346}.txt
    (widget SurveyJS; opsi muncul saat dropdown diklik → ditangkap flag --dump-opsi)

PAKAI:
  venv\\Scripts\\python.exe tools\\buat_pemetaan_full.py
"""
import glob
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(ROOT, "data", "output")
DUMP_FIELD = os.path.join(OUT_DIR, "diag_pelayanan_20260615_213912.txt")
OUT = os.path.join(OUT_DIR, "PEMETAAN_PELAYANAN_FULL.xlsx")
PEMETAAN_LAMA = os.path.join(OUT_DIR, "PEMETAAN_PELAYANAN.xlsx")

# baris INPUT: - 'label' | id='..' name='..' type='..' checked=.. value='..' ph='..'
RE_INPUT = re.compile(
    r"-\s*'(?P<label>.*?)'\s*\|\s*id='(?P<id>[^']*)'\s*name='(?P<name>[^']*)'\s*"
    r"type='(?P<type>[^']*)'\s*checked=(?P<checked>\w+)\s*value='(?P<value>[^']*)'\s*"
    r"ph='(?P<ph>[^']*)'"
)
RE_SQ = re.compile(r"sq_(\d+)i(?:_(\d+))?")
RE_PPM = re.compile(r"(PPM\d+)")


def _split_forms(text, header_re):
    """Pisah dump per form: kembalikan list (nama, isi_blok)."""
    parts = re.split(header_re, text)
    out = []
    # parts: [pre, nama1, blok1, nama2, blok2, ...]
    for i in range(1, len(parts), 2):
        out.append((parts[i], parts[i + 1]))
    return out


def parse_field_dump(path):
    """{form: {'questions': [ {sq,tipe,label,opsi:[(lbl,val)],ppm,id_base,ph} ],
                'headings': [..] }}"""
    text = open(path, encoding="utf-8").read()
    forms = {}
    for nama, blok in _split_forms(text, r"\[DIAG\] \(\d+/\d+\) Buka form: '([^']+)'"):
        # ambil bagian INPUT
        qmap = {}        # sq(int) -> question dict
        order = []
        for m in RE_INPUT.finditer(blok):
            d = m.groupdict()
            if d["type"] == "button":            # tombol Kirim
                continue
            sm = RE_SQ.search(d["id"])
            if not sm:
                continue
            sq = int(sm.group(1))
            if sq not in qmap:
                qmap[sq] = {"sq": sq, "tipe": None, "label": "", "judul": "",
                            "opsi": [], "ppm": "", "id_base": f"sq_{sq}i", "ph": d["ph"]}
                order.append(sq)
            q = qmap[sq]
            pm = RE_PPM.search(d["name"])
            if pm:
                q["ppm"] = pm.group(1)
            if d["type"] == "radio":
                q["tipe"] = "radio"
                q["opsi"].append((d["label"], d["value"]))
            elif d["type"] == "number":
                q["tipe"] = "number"
                q["label"] = d["label"]
            elif d["type"] == "text":
                q["tipe"] = "dropdown"
                q["ph"] = d["ph"]
        # QTITLE: sq -> judul pertanyaan (akurat, dari SurveyJS)
        qt = re.search(r"\] QTITLE \(sq -> judul pertanyaan\):\n(.*?)\n\[", blok, re.S)
        if qt:
            for ln in qt.group(1).splitlines():
                tm = re.match(r"\s*-\s*sq_(\d+):\s*'(.*)'\s*$", ln)
                if tm and int(tm.group(1)) in qmap:
                    qmap[int(tm.group(1))]["judul"] = tm.group(2)
        forms[nama] = {"questions": [qmap[s] for s in order]}
    return forms


def parse_opsi_dumps():
    """{form_partial: {sq(int): [opsi..]}} dari semua dump ber-[OPSI]."""
    res = {}
    for f in glob.glob(os.path.join(OUT_DIR, "diag_pelayanan_*.txt")):
        text = open(f, encoding="utf-8").read()
        fm = re.search(r"Membuka form pemeriksaan: '([^']+)'", text)
        if not fm or "[OPSI] id=" not in text:
            continue
        form = fm.group(1)
        d = res.setdefault(form, {})
        # blok: [OPSI] id='sq_100i_0' ...\n   opsi: [..]
        for m in re.finditer(r"\[OPSI\] id='sq_(\d+)i[^']*'.*?\n\s*opsi:\s*(\[[^\]]*\])",
                             text, re.S):
            sq = int(m.group(1))
            try:
                opsi = eval(m.group(2))          # list literal aman dari output kita
            except Exception:
                opsi = []
            if opsi:
                d[sq] = opsi
    return res


def cocok_opsi(form_name, opsi_map):
    """Cari entri opsi dump yg cocok dgn nama form (substring dua arah)."""
    for k, v in opsi_map.items():
        if k in form_name or form_name in k or k.split("(")[0].strip() in form_name:
            return v
    return {}


def baca_sumber():
    """{form: sumber} dari PEMETAAN lama (kolom Otomasi/Sumber)."""
    out = {}
    if not os.path.exists(PEMETAAN_LAMA):
        return out
    wb = openpyxl.load_workbook(PEMETAAN_LAMA, data_only=True)
    ws = wb["Daftar Form"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        nama, otomasi, sumber = r[3], r[4], r[5]
        if nama:
            out[str(nama).strip()] = (str(otomasi or "").strip(), str(sumber or "").strip())
    wb.close()
    return out


HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="2E7D32")
FILL_OV = PatternFill("solid", fgColor="4472C4")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)


def _style_header(ws, ncol, fill):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _safe_title(name, used):
    t = re.sub(r"[\\/?*\[\]:]", "-", name)[:31]
    base = t
    i = 2
    while t.lower() in used:
        t = f"{base[:28]}~{i}"
        i += 1
    used.add(t.lower())
    return t


def main():
    forms = parse_field_dump(DUMP_FIELD)
    opsi_dumps = parse_opsi_dumps()
    sumber = baca_sumber()

    wb = openpyxl.Workbook()
    ov = wb.active
    ov.title = "Daftar Form"
    ov.append(["No", "Form", "Otomasi?", "Sumber", "#Pertanyaan", "Sheet"])
    _style_header(ov, 6, FILL_OV)

    used = set()
    rownames = []
    for i, (nama, info) in enumerate(forms.items(), 1):
        rownames.append((i, nama, info))

    for i, nama, info in rownames:
        qs = info["questions"]
        # gabungkan opsi dropdown dari dump opsi
        ddmap = cocok_opsi(nama, opsi_dumps)
        for q in qs:
            if q["tipe"] == "dropdown" and q["sq"] in ddmap:
                q["opsi"] = [(o, "") for o in ddmap[q["sq"]]]
        sheet = _safe_title(nama, used)
        ws = wb.create_sheet(sheet)
        cols = ["No", "sq", "Tipe", "Pertanyaan", "Opsi",
                "Nilai Default (ISI)", "id base / PPM", "Catatan"]
        ws.append(cols)
        _style_header(ws, len(cols), FILL)
        for j, q in enumerate(qs, 1):
            opsi_txt = "  |  ".join(o[0] for o in q["opsi"])
            pert = q["judul"] or q["label"] or ""
            cat = ""
            if q["tipe"] == "dropdown" and not q["opsi"]:
                cat = "opsi belum tertangkap"
            elif not pert:
                cat = "judul tak tertangkap (cek form)"
            ws.append([j, q["sq"], q["tipe"], pert, opsi_txt, "",
                       f"{q['id_base']} / {q['ppm']}", cat])
        widths = [5, 7, 10, 38, 50, 22, 26, 30]
        for k, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + k)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = WRAP
                cell.border = THIN

        otom, sbr = sumber.get(nama, ("", ""))
        ov.append([i, nama, otom, sbr, len(qs), sheet])

    for k, w in enumerate([5, 52, 10, 12, 12, 32], 1):
        ov.column_dimensions[chr(64 + k)].width = w
    for row in ov.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN

    wb.save(OUT)
    tot = sum(len(v["questions"]) for v in forms.values())
    print(f"[OK] {OUT}")
    print(f"     {len(forms)} form, {tot} pertanyaan total.")
    # ringkas yg perlu perhatian
    for nama, info in forms.items():
        dd_kosong = [q["sq"] for q in info["questions"]
                     if q["tipe"] == "dropdown" and not cocok_opsi(nama, opsi_dumps).get(q["sq"])]
        if dd_kosong:
            print(f"     ! {nama}: dropdown tanpa opsi sq={dd_kosong}")


if __name__ == "__main__":
    main()
