"""
Tambah kolom INPUT data klinis ke Excel peserta agar tools/pelayanan.py bisa
mengisi field angka (Gizi, Gula Darah, Tekanan Darah). Header HARUS sama persis
dgn map EXCEL_NUM di pelayanan.py. Idempoten: kolom yg sudah ada dilewati.

PAKAI:
  venv\\Scripts\\python.exe tools\\tambah_kolom_klinis.py --excel data\\input\\template_pendaftaran.xlsx
"""
import argparse
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# (header, lebar, catatan utk komentar). Urutan = urutan form di pelayanan.
KOLOM = [
    ("Berat Badan", 12, "Gizi - kg (mis. 55.5)"),
    ("Tinggi Badan", 12, "Gizi - cm"),
    ("Lingkar Perut", 13, "Gizi - cm"),
    ("GDS", 8, "Gula Darah Sewaktu (mg/dl)"),
    ("GDS 2", 8, "GDS ke-2 (opsional, bila GDS1 tinggi)"),
    ("GDP", 8, "Gula Darah Puasa (mg/dl)"),
    ("GD2PP", 9, "Gula Darah 2 Jam PP (mg/dl)"),
    ("Sistolik", 9, "Tekanan Darah Sistolik"),
    ("Diastolik", 9, "Tekanan Darah Diastolik"),
    ("Sistolik 2", 10, "Sistolik ke-2 (opsional)"),
    ("Diastolik 2", 11, "Diastolik ke-2 (opsional)"),
]

HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="C0504D")    # merah-bata: tandai kolom INPUT klinis


def main():
    ap = argparse.ArgumentParser(description="Tambah kolom input klinis ke Excel peserta.")
    ap.add_argument("--excel", required=True)
    ap.add_argument("--header-row", dest="header_row", type=int, default=0)
    args = ap.parse_args()

    try:
        wb = openpyxl.load_workbook(args.excel)
    except PermissionError:
        sys.exit("File Excel sedang dibuka. TUTUP Excel lalu jalankan lagi.")
    ws = wb.worksheets[0]
    hdr = args.header_row + 1

    ada = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        if v is not None:
            ada[str(v).strip()] = c

    tambah = 0
    for nama, lebar, catatan in KOLOM:
        if nama in ada:
            print(f"  - '{nama}' sudah ada (kolom {ada[nama]}), lewati.")
            continue
        c = ws.max_column + 1
        cell = ws.cell(row=hdr, column=c, value=nama)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        try:
            cell.comment = openpyxl.comments.Comment(catatan, "pelayanan.py")
        except Exception:
            pass
        ws.column_dimensions[cell.column_letter].width = lebar
        print(f"  + '{nama}' ditambah di kolom {c} ({catatan}).")
        tambah += 1

    try:
        wb.save(args.excel)
    except PermissionError:
        sys.exit("Gagal simpan: file Excel sedang dibuka. TUTUP Excel lalu jalankan lagi.")
    print(f"[OK] {tambah} kolom baru ditambahkan ke {args.excel}.")


if __name__ == "__main__":
    main()
