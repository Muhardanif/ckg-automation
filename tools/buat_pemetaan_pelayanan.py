"""
Generator PEMETAAN_PELAYANAN.xlsx — dokumen kerja pemetaan form Step 3 (Pelayanan).

Sumber: enumerasi form dari tools/diag_pelayanan.py untuk peserta LANSIA PEREMPUAN
(SUMIATI). Dua sheet:
  1) 'Daftar Form'  : semua form yg muncul di detail-pemeriksaan, dikategorikan
     (Anamnesis 0/9 vs Pemeriksaan klinis). Kolom 'Otomasi?' & 'Sumber nilai'
     DIISI USER untuk menentukan cakupan (form mana yg akan diotomasi).
  2) 'Demografi Lansia' : contoh pemetaan field per-form (sudah terisi dari dump),
     jadi template untuk form-form lain yg akan kita dump berikutnya.

Jalankan ulang kapan saja (menimpa file). PAKAI:
  venv\\Scripts\\python.exe tools\\buat_pemetaan_pelayanan.py
"""
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "data", "output", "PEMETAAN_PELAYANAN.xlsx")

# (kategori, nama_form) — urutan sesuai dump diag_pelayanan (lansia perempuan).
ANAMNESIS = [
    "Demografi Lansia",
    "Faktor Risiko Kanker Usus",
    "Faktor Risiko TB - Dewasa & Lansia",
    "Hati",
    "Kanker Leher Rahim",
    "Kesehatan Jiwa",
    "Penapisan Risiko Kanker Paru",
    "Perilaku Merokok",
    "Tingkat Aktivitas Fisik (sedang dan berat)",
]
PEMERIKSAAN = [
    "Gizi (BB - TB - Lingkar Perut) Perempuan",
    "Pemeriksaan Gula Darah Dewasa Lansia",
    "Tekanan Darah Dewasa Lansia",
    "1. SKILAS Penurunan Kognitif - Lansia",
    "2. SKILAS Mobilisasi - Lansia",
    "3. SKILAS Malnutrisi",
    "4. SKILAS Pemeriksaan Gejala Depresi - Lansia",
    "5. Pemeriksaan Gangguan Fungsional/Barthel Index - Lansia",
    "6a. Penurunan Kognitif - Tindak Lanjut (Mini Cog-Clock Draw)",
    "6b. Penurunan Kognitif - Tindak Lanjut (AD-8 INA)",
    "7. Mobilisasi - Pemeriksaan Lanjutan (SPPB)",
    "8. Skrining Malnutrisi - Pemeriksaan Lanjutan (MNA-SF)",
    "9. Pemeriksaan Gejala Depresi - Pemeriksaan Lanjutan",
    "Faktor Risiko dan Skrining X-Ray TB (Dewasa & Lansia)",
    "Pemeriksaan Tuberkulosis (Dewasa & Lansia)",
    "Pemeriksaan Penyakit Frambusia",
    "Pemeriksaan Penyakit Kusta",
    "Pemeriksaan Penyakit Skabies",
    "Skrining Telinga dan Mata (=>40 tahun)",
    "Skrining Karies dan Gigi Hilang",
    "Skrining Penyakit Periodontal",
    "Pemeriksaan PPOK (Skrining PUMA)",
    "Pemeriksaan Kadar CO (merokok / terpapar asap)",
    "POCT Lipid Panel (>=40 thn & HT/DM)",
    "Pemeriksaan Fibrosis/Sirosis Hati",
    "Pemeriksaan Hepatitis",
    "Skrining Fungsi Ginjal Perempuan (>=40 thn risiko HT/DM)",
    "Skrining Kerusakan Ginjal (>=40 thn risiko HT/DM)",
    "Hasil Pemeriksaan - Skrining Jantung",
    "Skrining Kanker Payudara",
    "Hasil Pemeriksaan HPV-DNA",
    "Pemeriksaan Inspekulo dan IVA",
    "Skrining Kanker Paru (Usia >=45 thn)",
    "Pemeriksaan Lanjutan Kanker Usus",
]

HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="4472C4")
FILL2 = PatternFill("solid", fgColor="2E7D32")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)


def _style_header(ws, ncol, fill=FILL):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HDR
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def sheet_daftar(wb):
    ws = wb.active
    ws.title = "Daftar Form"
    cols = ["No", "Kelompok", "Kategori", "Nama Form",
            "Otomasi? (Ya/Tidak)", "Sumber nilai (default/excel/registrasi)",
            "Catatan"]
    ws.append(cols)
    _style_header(ws, len(cols))
    no = 1
    for kat, daftar in (("Anamnesis (hitung 0/9)", ANAMNESIS),
                        ("Pemeriksaan klinis", PEMERIKSAAN)):
        for nama in daftar:
            ws.append([no, "Lansia P", kat, nama, "", "", ""])
            no += 1
    widths = [5, 10, 22, 52, 18, 32, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN


def sheet_field_template(wb, judul, fields):
    """Sheet pemetaan field per-form. `fields` = list of dict."""
    ws = wb.create_sheet(judul[:31])
    cols = ["Form", "No.Field", "Field (label portal)", "Tipe", "Opsi",
            "Wajib?", "Nilai Default", "Kolom Excel", "id / selector", "Catatan"]
    ws.append(cols)
    _style_header(ws, len(cols), fill=FILL2)
    for i, f in enumerate(fields, 1):
        ws.append([judul, i, f.get("label", ""), f.get("tipe", ""),
                   f.get("opsi", ""), f.get("wajib", ""), f.get("default", ""),
                   f.get("kolom", ""), f.get("sel", ""), f.get("catatan", "")])
    widths = [18, 8, 30, 10, 34, 12, 18, 18, 30, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = THIN


def main():
    wb = openpyxl.Workbook()
    sheet_daftar(wb)

    # contoh pemetaan terisi: Demografi Lansia (dari dump form.kemkes.go.id)
    demografi = [
        {"label": "Status Perkawinan", "tipe": "radio",
         "opsi": "Belum Menikah / Menikah / Cerai Mati / Cerai Hidup",
         "wajib": "wajib (*)", "default": "(dari registrasi)",
         "kolom": "status_pernikahan", "sel": "id=sq_100i_0..3 / name=...PPM00000172",
         "catatan": "pakai ulang data pendaftaran 'Status Pernikahan'"},
        {"label": "Disabilitas", "tipe": "radio",
         "opsi": "Non disabilitas / Penyandang disabilitas",
         "wajib": "?", "default": "Non disabilitas",
         "kolom": "disabilitas", "sel": "id=sq_101i_0..1 / name=...PPM00000299",
         "catatan": "default 'Non disabilitas'; override dari registrasi bila ada"},
    ]
    sheet_field_template(wb, "Demografi Lansia", demografi)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"[OK] PEMETAAN dibuat: {OUT}")
    print(f"     Daftar Form: {len(ANAMNESIS)} anamnesis + {len(PEMERIKSAAN)} pemeriksaan.")


if __name__ == "__main__":
    main()
