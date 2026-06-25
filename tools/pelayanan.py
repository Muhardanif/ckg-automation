"""
PELAYANAN (Step 3 CKG) — isi 22 form skrining per peserta secara otomatis (CDP).

Konfigurasi jawaban dibaca SAAT RUN dari workbook pemetaan
`data/output/PEMETAAN_PELAYANAN_FULL.xlsx` (kolom 'Nilai Default (ISI)'), jadi
hasil review user langsung berlaku tanpa ubah kode. Nilai per-peserta (angka
klinis & Status Perkawinan) diambil dari Excel peserta bila kolomnya tersedia.

ALUR per peserta (mengikuti pola yg terbukti di tools/diag_pelayanan.py):
  listing -> filter NIK (tanpa tanggal; fallback filter Nama + tanggal) -> cari
  -> cocokkan baris via NAMA -> klik 'Mulai' -> detail-pemeriksaan
  -> (opsional) 'Mulai Pemeriksaan' + modal tanggal (tulis tgl ke Excel)
  -> untuk tiap form Otomasi=Ya: buka 'Input Data' -> isi field -> (Kirim)
     -> goto detail_url (BUKAN tombol 'Kembali' yg sering balik ke detail kosong)
  -> (opsional) 'Selesaikan Layanan'.

KESELAMATAN: default DRY-RUN — form DIISI tapi TIDAK di-Kirim (tak mengubah
server). Pakai --submit utk benar-benar mengirim tiap form, --selesaikan utk
klik 'Selesaikan Layanan'. Uji dulu dgn --nik <satu peserta> --dry-run.

PERSIAPAN: Chrome --remote-debugging-port=9222 (1_mulai_chrome.bat), login,
buka CKG Umum. TUTUP Excel (skrip menulis balik status).

PAKAI:
  venv\\Scripts\\python.exe tools\\pelayanan.py --excel data\\input\\template_pendaftaran.xlsx --nik 3525084710630003 --tab "Sedang Pemeriksaan"
  # tambah --submit --selesaikan utk produksi; --forms "Hati,Demografi Lansia" utk batasi
"""
import argparse
import asyncio
import io
import os
import re
import sys
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl                                                  # noqa: E402

from app.automation.ckg_bot import CKGBot                        # noqa: E402
from app.automation import selectors as S                        # noqa: E402
from app.schema import KelompokUsia                              # noqa: E402
from app.readers import baca_excel                               # noqa: E402
from app.excel_hasil import (KOL_WAKTU_HADIR, KOL_STATUS_HADIR,  # noqa: E402
                             STATUS_HADIR_TERMINAL)

import diag_pelayanan as dp                                      # noqa: E402 (modul tetangga)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PEMETAAN = os.path.join(ROOT, "data", "output", "PEMETAAN_PELAYANAN_FULL.xlsx")
URL_PELAYANAN = dp.URL_PELAYANAN

# Kolom hasil tahap pelayanan (ditulis balik ke Excel peserta).
KOL_STATUS_LAYANAN = "Status Layanan"
KOL_WAKTU_LAYANAN = "Waktu Layanan"
KOL_TGL_PERIKSA = "Tanggal Pemeriksaan"
KOL_WAKTU_MULAI = "Waktu Mulai Periksa"      # timestamp saat klik 'Mulai Pemeriksaan'
KOL_WAKTU_SELESAI = "Waktu Selesai Periksa"  # timestamp saat 'Selesaikan Layanan' sukses
STATUS_LAYANAN_OK = "SELESAI"
STATUS_LAYANAN_DRY = "DRAFT (dry-run)"
# Transaksi pelayanan sudah selesai di portal (Pemeriksaan Mandiri 'Lengkap' +
# Pelayanan 'Selesai Pemeriksaan') → terminal: dicatat & dilewati, tak diproses.
STATUS_SUDAH_SELESAI = "SUDAH SELESAI"

EXCEL = "(dari Excel/alat)"

# Field angka per-peserta -> nama kolom Excel (dibuat user bila ingin diisi).
# key = (substring nama form, sq). Bila kolom tak ada / kosong -> field dilewati.
EXCEL_NUM = {
    ("Gizi (BB", 100): "Berat Badan", ("Gizi (BB", 101): "Tinggi Badan",
    ("Gizi (BB", 102): "Lingkar Perut",
    ("Gula Darah", 102): "GDS", ("Gula Darah", 103): "GDS 2",
    ("Gula Darah", 104): "GDP", ("Gula Darah", 105): "GD2PP",
    ("Tekanan Darah", 102): "Sistolik", ("Tekanan Darah", 103): "Diastolik",
    ("Tekanan Darah", 104): "Sistolik 2", ("Tekanan Darah", 105): "Diastolik 2",
    # Set REMAJA / Anak Sekolah (nama form & sq berbeda dari dewasa).
    ("Gizi Anak Sekolah", 100): "Berat Badan", ("Gizi Anak Sekolah", 101): "Tinggi Badan",
    ("Tekanan Darah Anak dan Remaja", 100): "Sistolik",
    ("Tekanan Darah Anak dan Remaja", 101): "Diastolik",
    ("Pemeriksaan Gula Darah Remaja", 102): "GDS",
    ("Pemeriksaan Gula Darah Remaja", 103): "GDS 2",
    ("Pemeriksaan Gula Darah Anak", 102): "GDS",
    ("Pemeriksaan Gula Darah Anak", 103): "GDS 2",
}
# Form WAJIB hijau/'Selesai diperiksa' sebelum boleh klik 'Selesaikan Layanan'
# (gating kunci). Cocokkan via prefix nama kartu ternormalisasi. Form yg ABSEN
# pada peserta (mis. 'demografi' versi lain) otomatis diabaikan.
WAJIB_SELESAIKAN = [
    "demografi",                       # Pemeriksaan Mandiri (lansia / dewasa P/L)
    "faktor risiko kanker usus", "faktor risiko tb", "hati", "kanker leher rahim",
    "kesehatan jiwa", "penapisan risiko kanker paru", "perilaku merokok",
    "tingkat aktivitas fisik",
    "gizi", "pemeriksaan gula darah", "tekanan darah",   # TB/BB/Gula/Tensi
    "skrining telinga dan mata", "skrining karies", "skrining penyakit periodontal",
    "pemeriksaan ppok",                                  # Telinga/Mata/Gigi/PPOK
]

# Kartu NAKES/lab yg SAH tetap 'Dalam Pemeriksaan' / belum selesai = di LUAR scope
# otomasi (diisi nakes lain / lab eksternal). Cocokkan via substring nama kartu
# ternormalisasi. Kartu NAKES yg BELUM selesai & di luar daftar ini = BLOCKER untuk
# 'Selesaikan Layanan' → jangan kunci data tak lengkap. (Sesuaikan bila perlu.)
# PENTING: LUAR_SCOPE TIDAK berlaku utk Pemeriksaan Mandiri (baris <tr>) — aturan
# user: SEMUA Mandiri wajib hijau (lihat _mandiri_belum). Ini hanya utk kartu Nakes.
# Daftar diverifikasi dari peserta dewasa-L yg ter-LOCK normal (ACHMAD, 2026-06-16):
# kartu² ini tetap 'Dalam Pemeriksaan' & portal TETAP izinkan kunci → out-of-scope.
LUAR_SCOPE = [
    "laboratorium", "tes iva", "pemeriksaan iva", "calon pengantin",
    "kadar co", "x-ray tb", "pemeriksaan tuberkulosis", "fibrosis",
    "pemeriksaan hepatitis", "hiv", "sifilis", "frambusia", "kusta", "skabies",
    "penyakit tropis",
    # tambahan dari set REMAJA (PUTRI): hasil kebugaran jasmani (diukur nakes) &
    # tes RDT malaria (lab). 'kebugaran jasmani' ≠ 'Kelayakan Tes Kebugaran'
    # (kuesioner in-scope); 'rdt malaria' ≠ 'Faktor Risiko Malaria' (anamnesis).
    "kebugaran jasmani", "rdt malaria",
    # tambahan dari set ANAK SD (AZKA): lab Hepatitis B usia 7-12 & TB anak.
    # 'hepatitis b usia' ≠ 'Faktor Risiko Hepatitis SD' (anamnesis in-scope).
    "hepatitis b usia",
    # lab/prosedur usia >=40 (MONALI 50th) — portal TIDAK mewajibkan (peserta tetap
    # bisa terkunci dgn ini 'Dalam Pemeriksaan'): POCT lipid, panel ginjal, lanjutan
    # kanker usus. 'lanjutan kanker usus' ≠ 'Faktor Risiko Kanker Usus' (anamnesis).
    "poct", "fungsi ginjal", "kerusakan ginjal", "lanjutan kanker usus",
    # lab/prosedur ginekologi perempuan (NUR): HPV-DNA & inspekulo IVA.
    "hpv", "inspekulo",
]

# Status Perkawinan: vocab Excel/Peserta -> opsi portal.
PERNIKAHAN = {
    "kawin": "Menikah", "menikah": "Menikah",
    "belum kawin": "Belum Menikah", "belum menikah": "Belum Menikah",
    "cerai hidup": "Cerai Hidup", "cerai mati": "Cerai Mati",
}


def log(msg):
    try:                                 # konsol Windows (cp1252) bisa gagal utk →, ≥, dll.
        print(f"[PELAYANAN] {msg}", flush=True)
    except UnicodeEncodeError:
        print(f"[PELAYANAN] {msg.encode('ascii', 'replace').decode('ascii')}", flush=True)


def baca_config(path):
    """Baca workbook pemetaan -> list (form_name, sheet, [pertanyaan...]).
    pertanyaan = {sq, tipe, opsi:[label..], default}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ov = wb["Daftar Form"]
    forms = []
    for r in ov.iter_rows(min_row=2, values_only=True):
        no, nama, otomasi, sumber, njml, sheet = r[:6]
        if not nama or sheet not in wb.sheetnames:
            continue
        qs = []
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            sq, tipe, judul, opsi, default = row[1], row[2], row[3], row[4], row[5]
            if sq is None:
                continue
            opts = [o.strip() for o in (opsi or "").split("|") if o.strip()]
            qs.append({"sq": int(sq), "tipe": tipe, "judul": judul or "",
                       "opsi": opts, "default": (default or "").strip()})
        forms.append({"nama": nama, "sheet": sheet, "qs": qs})
    wb.close()
    return forms


def _excel_num_cols(ws, hdr):
    """{header: col_idx} utk kolom angka klinis yg ADA di Excel."""
    want = set(EXCEL_NUM.values())
    found = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        h = str(v).strip() if v is not None else ""
        if h in want:
            found[h] = c
    return found


async def _isi_radio(page, sq, idx):
    """Pilih radio ke-idx pada pertanyaan sq (klik label, fallback check)."""
    lab = page.locator(f"label[for='sq_{sq}i_{idx}']")
    if await lab.count() > 0:
        await lab.first.click(timeout=5000)
        return True
    await page.locator(f"#sq_{sq}i_{idx}").check(timeout=5000, force=True)
    return True


async def _isi_dropdown(page, sq, teks):
    """Buka dropdown SurveyJS (klik wrapper input sq_{sq}i_0) lalu klik opsi `teks`.

    Opsi dibatasi ke listbox milik pertanyaan ini (`ul#sq_{sq}i_list`) — penting
    karena banyak dropdown berbagi label opsi sama (mis. 'Mandiri') sehingga
    pencarian global salah sasaran ke popup lain yg tersembunyi."""
    inp = page.locator(f"#sq_{sq}i_0")
    lst = f"ul#sq_{sq}i_list li[role='option']"
    await page.keyboard.press("Escape")              # tutup popup dropdown sebelumnya
    await page.wait_for_timeout(250)
    try:                                             # popup render dekat input; bawa ke viewport
        await inp.scroll_into_view_if_needed(timeout=3000)
    except Exception:
        pass
    dibuka = False
    for attempt in range(3):                         # retry buka bila popup belum muncul
        for xp in ("xpath=ancestor::div[1]", "xpath=ancestor::div[2]", "xpath=ancestor::div[3]"):
            try:
                await inp.locator(xp).first.click(timeout=3000)
            except Exception:
                continue
            for _ in range(8):                       # polling ~2.4s sampai listbox render
                if await page.locator(lst).count() > 0:
                    dibuka = True
                    break
                await page.wait_for_timeout(300)
            if dibuka:
                break
        if dibuka:
            break
        await page.wait_for_timeout(300)
    opt = page.locator(lst).filter(has_text=re.compile(rf"^\s*{re.escape(teks)}\s*$"))
    if await opt.count() == 0:                       # fallback: cocokkan longgar
        opt = page.locator(lst).filter(has_text=teks)
    await opt.first.click(timeout=5000)
    await page.wait_for_timeout(300)


async def _isi_number(page, sq, val):
    await page.locator(f"#sq_{sq}i").fill(str(val), timeout=5000)


async def _sudah_terisi(page, sq, tipe):
    """True bila pertanyaan sq sudah punya jawaban (utk mode --resume → jangan timpa)."""
    try:
        if tipe == "radio":
            return await page.locator(f"input[id^='sq_{sq}i_']:checked").count() > 0
        if tipe == "number":
            v = await page.locator(f"#sq_{sq}i").input_value(timeout=2000)
            return bool((v or "").strip())
        if tipe == "dropdown":
            cls = await page.locator(f"#sq_{sq}i").get_attribute("class") or ""
            return "sd-dropdown--empty" not in cls
    except Exception:
        return False
    return False


# JS: status tiap kartu form di halaman detail (tanpa membuka form).
# - Pemeriksaan Mandiri (baris <tr>): selesai bila ada centang HIJAU
#   (icon-success.svg). PENTING: centang ABU-ABU (belum) = 'icon-success-GRAY.svg'
#   yg JUGA memuat substring 'icon-success' → WAJIB kecualikan varian 'gray',
#   kalau tidak form belum-terisi terbaca selesai (false-positive).
# - Pelayanan Nakes (kartu non-tr): selesai bila teks memuat 'Selesai diperiksa'.
JS_STATUS_KARTU = r"""()=>{
 const txt=e=>(e.textContent||'').replace(/\s+/g,' ').trim();
 const res=[];
 document.querySelectorAll('tr').forEach(tr=>{
   if(!/input data/i.test(txt(tr))) return;
   const td=tr.querySelector('td'); if(!td) return;
   res.push({nama: txt(td),
             done: !!tr.querySelector("img[src*='icon-success']:not([src*='gray'])")});
 });
 document.querySelectorAll('button').forEach(b=>{
   if(!/input data/i.test(txt(b))||b.closest('tr')) return;
   let card=b, t='';
   for(let i=0;i<7&&card;i++){card=card.parentElement; if(!card)break;
     const s=txt(card); if(s.length>10&&s.length<200){t=s;break;}}
   if(!t) return;
   const nama=t.replace(/\bYa\b/,'').replace(/Selesai diperiksa.*$/i,'')
              .replace(/Dalam Pemeriksaan.*$/i,'').replace(/Input Data.*$/i,'').trim();
   res.push({nama, done: /Selesai diperiksa/i.test(t)});
 });
 return res;
}"""


async def _peta_status(page):
    """Baca status semua kartu form di detail → list (nama_lower, done)."""
    try:
        rows = await page.evaluate(JS_STATUS_KARTU)
    except Exception:
        return []
    return [((r.get("nama") or "").lower(), bool(r.get("done"))) for r in rows]


# JS: kartu PEMERIKSAAN MANDIRI (baris <tr>) yg BELUM hijau (centang non-gray).
JS_MANDIRI_BELUM = r"""()=>{
 const txt=e=>(e.textContent||'').replace(/\s+/g,' ').trim();
 const res=[];
 document.querySelectorAll('tr').forEach(tr=>{
   if(!/input data/i.test(txt(tr))) return;
   const td=tr.querySelector('td'); if(!td) return;
   const done=!!tr.querySelector("img[src*='icon-success']:not([src*='gray'])");
   if(!done) res.push(txt(td).slice(0,60));
 });
 return res;
}"""


async def _mandiri_belum(page):
    """Daftar kartu PEMERIKSAAN MANDIRI (baris <tr>) yg BELUM hijau.

    ATURAN user (2026-06-16): SEMUA Pemeriksaan Mandiri WAJIB hijau sebelum
    'Selesaikan Layanan' — TANPA pengecualian. LUAR_SCOPE hanya berlaku utk kartu
    Nakes/lab (non-tr), TIDAK utk Mandiri. Jadi Mandiri dicek terpisah & ketat."""
    try:
        return [s for s in (await page.evaluate(JS_MANDIRI_BELUM)) if s]
    except Exception:
        return []


def _norm_kartu(s):
    """Normalisasi nama kartu/judul: buang nomor awal ('1. '/'6a. ') & ekor 'ya'."""
    s = re.sub(r"^\s*\d+[a-z]?\.\s*", "", s or "")
    s = re.sub(r"\s*ya$", "", s.strip(), flags=re.I)
    return s.strip().lower()


def _frag(judul):
    """Potongan judul aman utk cocokkan kartu (buang nomor awal & kurung)."""
    s = re.sub(r"^\s*\d+[a-z]?\.\s*", "", judul)
    s = s.split("(")[0].split("=>")[0].split(">=")[0].strip()
    return s[:40].lower()


def _wajib_belum_selesai(status_map):
    """Daftar form WAJIB yg ADA di detail tapi BELUM hijau/'Selesai diperiksa'.
    Form wajib yg absen pd peserta diabaikan. Non-kosong → JANGAN klik Selesaikan."""
    belum = []
    for frag in WAJIB_SELESAIKAN:
        matches = [done for nama, done in status_map if _norm_kartu(nama).startswith(frag)]
        if matches and not all(matches):
            belum.append(frag)
    return belum


def _kartu_belum_diluar_izin(status_map):
    """SEMUA kartu di detail yg BELUM selesai & BUKAN out-of-scope (LUAR_SCOPE).
    Ini PENGAMAN utama: non-kosong → ada form belum lengkap (termasuk form yg
    TAK ADA di config / belum dipetakan) → JANGAN kunci ('Selesaikan Layanan').
    Lebih ketat dari _wajib_belum_selesai (yg hanya cek daftar WAJIB)."""
    blk = []
    for nama, done in status_map:
        if done:
            continue
        n = _norm_kartu(nama)
        if any(s in n for s in LUAR_SCOPE):
            continue
        blk.append((nama or "").strip())
    return blk


def _kartu_tak_dikenal(status_map, forms):
    """Kartu di detail yg BELUM selesai, di luar out-of-scope, & TAK cocok satu pun
    form di config (= form belum dipetakan; mis. Demografi Dewasa Laki-Laki,
    Imunisasi Tetanus Catin, form remaja). Hanya utk PERINGATAN/visibilitas."""
    frags = [_frag(f["nama"]) for f in forms]
    out = []
    for nama in _kartu_belum_diluar_izin(status_map):
        n = _norm_kartu(nama)
        if not any(n.startswith(fr) for fr in frags if fr):
            out.append(nama)
    return out


def _forms_aktif(forms, status_map):
    """Untuk tiap kartu di detail, pilih config form dgn frag TERPANJANG yg cocok
    (prefix nama kartu ternormalisasi). Kembalikan set nama form yg 'menang' di
    minimal satu kartu. Mencegah form generik (mis. 'Perilaku Merokok',
    'Gizi (BB..)', 'Skrining Telinga dan Mata') ikut terpicu di kartu varian
    lebih spesifik ('Perilaku Merokok - Anak Sekolah', 'Gizi Anak Sekolah',
    'Skrining Telinga dan Mata - Anak Sekolah') → tiap kartu diisi 1 config saja
    sesuai kelompok usia peserta."""
    aktif = set()
    for nama_kartu, _done in status_map:
        nk = _norm_kartu(nama_kartu)
        best, best_len = None, -1
        for f in forms:
            fr = _frag(f["nama"])
            if fr and nk.startswith(fr) and len(fr) > best_len:
                best, best_len = f["nama"], len(fr)
        if best:
            aktif.add(best)
    return aktif


def _form_selesai(status_map, judul):
    """True bila kartu form `judul` berstatus selesai di detail (None bila tak ketemu).
    Cocokkan via nama kartu ternormalisasi yg DIAWALI frag judul."""
    frag = _frag(judul)
    cocok = [done for nama, done in status_map if _norm_kartu(nama).startswith(frag)]
    if not cocok:
        return None
    return all(cocok)


async def _visible_sq(page):
    """Set sq (int) pertanyaan yg TERLIHAT sekarang (deteksi munculnya kondisional)."""
    form = await page.evaluate(dp.JS_DUMP_FORM)
    out = set()
    for i in form["inputs"]:
        if i["type"] == "button":
            continue
        m = re.search(r"sq_(\d+)i", i["id"])
        if m:
            out.add(int(m.group(1)))
    return out


async def _isi_pertanyaan(page, form_nama, q, pvals):
    """Isi satu pertanyaan. Return (ok, pesan_bila_lewat|None)."""
    sq, tipe, opsi, default = q["sq"], q["tipe"], q["opsi"], q["default"]
    if tipe == "number":
        val = pvals.get(("num", form_nama, sq))
        if val in (None, "") and default and default != EXCEL and str(default).strip():
            val = default                              # default literal angka (mis. kondisional)
        if val in (None, ""):
            return False, f"sq{sq} angka kosong (lewati)"
        await _isi_number(page, sq, val)
    elif tipe == "radio":
        pilih = pvals.get(("perkawinan",)) or "" if default == EXCEL else default
        if not pilih or pilih not in opsi:
            return False, f"sq{sq} radio default {default!r} tak cocok opsi"
        await _isi_radio(page, sq, opsi.index(pilih))
    elif tipe == "dropdown":
        if not default or default not in opsi:
            return False, f"sq{sq} dropdown default {default!r} tak cocok opsi"
        await _isi_dropdown(page, sq, default)
    else:
        return False, f"sq{sq} tipe {tipe!r} tak didukung"
    await page.wait_for_timeout(150)
    return True, None


async def isi_satu_form(page, form, pvals, dry, log, resume=False):
    """Isi semua pertanyaan satu form secara ITERATIF: isi yg terlihat, re-scan,
    isi pertanyaan kondisional yg baru muncul, ulang sampai stabil. pvals = nilai
    per-peserta.

    resume=True: pertanyaan yg SUDAH terisi DILEWATI (tak ditimpa); form yg tak
    ada isian baru TIDAK di-Kirim ulang. Return (n_isi, n_lewat, n_skip, catatan[])."""
    cfg = {q["sq"]: q for q in form["qs"]}
    filled = set()
    n_isi = n_lewat = n_skip = 0
    cat = []
    for _ in range(8):                                  # cukup utk beberapa lapis kondisional
        visible = await _visible_sq(page)
        todo = sorted(sq for sq in cfg if sq in visible and sq not in filled)
        if not todo:
            break
        for sq in todo:
            filled.add(sq)
            if resume and await _sudah_terisi(page, sq, cfg[sq]["tipe"]):
                n_skip += 1
                continue
            try:
                ok, pesan = await _isi_pertanyaan(page, form["nama"], cfg[sq], pvals)
            except Exception as e:
                ok, pesan = False, f"sq{sq} ERROR: {type(e).__name__}: {str(e)[:70]}"
            if ok:
                n_isi += 1
            else:
                n_lewat += 1
                cat.append(pesan)
    # deteksi pertanyaan kondisional yg TERLIHAT tapi tak ada di config
    visible = await _visible_sq(page)
    unmapped = sorted(sq for sq in visible if sq not in cfg)
    if unmapped:
        cat.append(f"PERLU MAP: kondisional sq={unmapped} (tak terisi)")
    # submit / dry-run. Resume: jangan Kirim bila tak ada isian baru (form sudah lengkap).
    if not dry and not (resume and n_isi == 0):
        kirim = page.locator("input[type='button'][value='Kirim'], button:has-text('Kirim')")
        try:
            await kirim.first.click(timeout=6000)
            await page.wait_for_timeout(1500)
        except Exception as e:
            cat.append(f"KIRIM gagal: {str(e)[:100]}")
    return n_isi, n_lewat, n_skip, cat


async def _konfirmasi_kunci(page, log):
    """Tangani dialog 'Data Pemeriksaan akan Dikunci' setelah 'Selesaikan Layanan':
    klik 'Konfirmasi' (BUKAN 'Periksa Kembali'). Return True bila terkonfirmasi.
    PENTING: aksi ini MENGUNCI data peserta (final, dikirim ke rapor WA/SATUSEHAT)."""
    await page.wait_for_timeout(900)
    # pastikan dialog kunci muncul (teks penanda), agar tak salah klik tombol lain
    dialog = page.get_by_text(re.compile(r"Data Pemeriksaan akan Dikunci|menjadi hasil akhir",
                                         re.I))
    try:
        if await dialog.count() == 0:
            await page.wait_for_timeout(1200)        # beri waktu render
    except Exception:
        pass
    konfirm = page.get_by_role("button", name=re.compile(r"^\s*Konfirmasi\s*$", re.I))
    if await konfirm.count() == 0:
        konfirm = page.get_by_text(re.compile(r"^\s*Konfirmasi\s*$", re.I)).locator("visible=true")
    if await konfirm.count() == 0:
        return False
    try:
        await konfirm.first.click(timeout=6000)
    except Exception as e:
        log(f"  gagal klik 'Konfirmasi': {str(e)[:80]}")
        return False
    # tunggu overlay dialog hilang
    try:
        await dp._tunggu_overlay_hilang(page)
    except Exception:
        pass
    await page.wait_for_timeout(1500)
    return True


async def _kembali_ke_detail(page, detail_url):
    """Navigasi LANGSUNG ke URL detail (andal) + tunggu kartu 'Input Data'."""
    try:
        await page.goto(detail_url)
        await page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        pass
    ib = page.get_by_role("button", name=re.compile(r"Input Data", re.I))
    for _ in range(20):
        if await ib.count() > 0:
            return True
        await page.wait_for_timeout(500)
    return False


def _now():
    return datetime.now().isoformat(timespec="seconds")


async def _cari_di_tabs(page, fh, bot, nama, tabs, aksi, tgl_filter, nik=None,
                        lewati_selesai=False):
    """Cari peserta lalu buka detail di salah satu tab kandidat.

    Mode NIK (bila `nik` diberikan): filter dropdown -> 'NIK', cari NIK. TANPA
    filter tanggal (temuan user 2026-06-16: filter NIK tak butuh input tanggal
    hadir -> lebih cepat). Baris TETAP dicocokkan via NAMA karena listing tak
    menampilkan kolom NIK; namun filter NIK menyisakan 1 baris unik (aman utk
    nama kembar).
    Mode Nama (fallback): filter tanggal (Waktu Hadir) -> filter 'Nama' -> cari.

    Filter+kata di-set SEKALI; lalu klik tiap tab kandidat sampai baris (cocok
    NAMA) muncul & bisa dibuka. Return nama tab tempat ketemu, atau None."""
    if nik:
        await dp._set_dropdown_filter(page, fh, "NIK")
        # poll pendek: peserta sering ada di tab LAIN → tak perlu tunggu lama di
        # tab aktif sekarang; pemilihan tab di bawah yg memastikan barisnya muncul.
        await dp._cari(page, fh, nik, match_text=nama, tries=4)
        kata = nik
        log(f"  filter NIK (tanpa filter tanggal) = {nik}")
    else:
        if tgl_filter:
            await dp._set_tanggal_rentang(bot, fh, tgl_filter)
            log(f"  filter tanggal listing = {tgl_filter}")
        await dp._set_dropdown_filter(page, fh, "Nama")
        await dp._cari(page, fh, nama, tries=4)
        kata = nama
    for tab in tabs:
        await dp._pilih_tab(page, fh, tab)
        await page.wait_for_timeout(700)
        # bila pindah tab mengosongkan pencarian, ulangi cari sekali (cocok via NAMA)
        if await page.locator("tbody tr").filter(has_text=nama).count() == 0:
            await dp._cari(page, fh, kata, match_text=nama, tries=6)
        # ATURAN user: bila baris menandakan transaksi SUDAH SELESAI (Pemeriksaan
        # Mandiri 'Lengkap' + Pelayanan 'Selesai Pemeriksaan'), JANGAN klik
        # 'Mulai' — cukup catat di log & lewati. (Tetap diproses saat dry-run /
        # debug agar bisa cek format.)
        if lewati_selesai and await dp._baris_transaksi_selesai(page, nama):
            log(f"  TRANSAKSI SUDAH SELESAI di tab '{tab}' (Pemeriksaan Mandiri "
                f"'Lengkap' & Pelayanan 'Selesai Pemeriksaan'). Tidak klik 'Mulai'.")
            return dp.SUDAH_SELESAI
        if await dp._klik_mulai(page, fh, nama, aksi=aksi):
            return tab
    return None


async def proses_peserta(page, fh, bot, p, forms, args, pvals, tabs, auto_mulai, tgl_filter):
    """Proses satu peserta. `tabs` = daftar tab listing yg dicoba berurutan (dari
    Excel). `auto_mulai`: True → klik 'Mulai Pemeriksaan' hanya bila peserta
    ditemukan di tab 'Belum Pemeriksaan'. `tgl_filter` = tanggal (ISO) utk filter
    listing (dari 'Waktu Hadir'). Return dict hasil."""
    res = {"status": None, "tgl_periksa": None, "waktu_mulai": None,
           "waktu_selesai": None, "mulai_ok": False, "selesai_ok": False, "tab": None}
    nama = (p.nama or "").strip()
    if not nama:
        res["status"] = "GAGAL: nama kosong"
        return res

    # 1) listing: set filter tanggal (Waktu Hadir) lalu coba tiap tab kandidat
    if not (page.url or "").rstrip("/").endswith("/ckg-pelayanan"):
        await page.goto(URL_PELAYANAN)
        try:
            await page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass
    await page.wait_for_timeout(1200)
    nik = (p.nik or "").strip()
    # 1a) JALUR UTAMA: filter NIK (tanpa filter tanggal -> lebih cepat).
    found_tab = None
    # Lewati klik 'Mulai' bila transaksi sudah selesai — KECUALI dry-run (cek
    # format/debug), yg tetap perlu membuka detail.
    lewati_selesai = not args.dry_run
    if nik and not args.filter_nama:
        found_tab = await _cari_di_tabs(page, fh, bot, nama, tabs, args.aksi, None,
                                        nik=nik, lewati_selesai=lewati_selesai)
        if found_tab is None:
            log("  (NIK) tak ketemu -> fallback ke filter Nama + tanggal")
            await page.goto(URL_PELAYANAN)
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
            await page.wait_for_timeout(1000)
    # 1b) FALLBACK: filter Nama + filter tanggal (Waktu Hadir).
    if found_tab is None:
        found_tab = await _cari_di_tabs(page, fh, bot, nama, tabs, args.aksi, tgl_filter,
                                        lewati_selesai=lewati_selesai)
    # fallback: bila tak ketemu dgn filter Waktu Hadir, ulangi TANPA filter tanggal
    # (mis. tgl pemeriksaan di listing beda dari Waktu Hadir).
    if found_tab is None and tgl_filter:
        log("  tak ketemu dgn filter Waktu Hadir -> coba lagi tanpa filter tanggal")
        await page.goto(URL_PELAYANAN)
        try:
            await page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass
        await page.wait_for_timeout(1000)
        found_tab = await _cari_di_tabs(page, fh, bot, nama, tabs, args.aksi, None,
                                        lewati_selesai=lewati_selesai)
    # Transaksi sudah selesai → JANGAN diproses ulang; cukup catat di log.
    if found_tab is dp.SUDAH_SELESAI:
        log("  -> Dilewati: transaksi pelayanan sudah selesai (tidak klik 'Mulai').")
        res["tab"] = "Selesai Pemeriksaan"
        res["status"] = STATUS_SUDAH_SELESAI
        return res
    if found_tab is None:
        res["status"] = (f"GAGAL: tak bisa buka detail (tak ada di listing ATAU sudah "
                         f"terkunci/tanpa tombol aksi; tab dicoba: {tabs})")
        return res
    res["tab"] = found_tab
    detail_url = page.url
    tgl_iso = None
    # klik 'Mulai Pemeriksaan' hanya bila peserta memang di tab 'Belum' (atau dipaksa)
    do_mulai = (found_tab == "Belum Pemeriksaan") if auto_mulai else args.mulai_pemeriksaan
    log(f"  ditemukan di tab '{found_tab}'; mulai-pemeriksaan={do_mulai}")

    # 2) mulai pemeriksaan (modal tanggal) bila perlu → catat tgl & waktu mulai
    if do_mulai:
        tombol_mulai = page.get_by_role("button", name=re.compile(r"^\s*Mulai Pemeriksaan\s*$", re.I))
        if await tombol_mulai.count() > 0:
            await dp._klik_teks(page, fh, "Mulai Pemeriksaan")
            tgl_iso = args.tanggal_periksa or _waktu_hadir_iso(p) or date.today().isoformat()
            ok_mulai = await dp._lewati_modal_mulai(page, fh, tgl_iso)
            await dp._tunggu_overlay_hilang(page)
            detail_url = page.url
            if ok_mulai and not args.dry_run:
                res["mulai_ok"] = True
                res["tgl_periksa"] = tgl_iso
                res["waktu_mulai"] = _now()
                log(f"  MULAI PEMERIKSAAN: tgl={tgl_iso} waktu={res['waktu_mulai']}")
        else:
            log("  (Mulai Pemeriksaan tak ada — peserta mungkin sudah dimulai.)")

    # 4) isi tiap form (pvals = nilai per-peserta sudah disiapkan pemanggil)
    tot_isi = tot_lewat = tot_skip = 0
    n_skip_form = n_skip_absen = 0
    # baca status kartu di detail SEKALI: utk (a) filter form yg TAK ADA pd peserta
    # ini (beda kelompok/JK, mis. SKILAS geriatri utk dewasa), (b) skip form selesai.
    status_map = await _peta_status(page)
    # config form mana yg BERLAKU utk peserta ini (1 kartu = config frag terpanjang).
    aktif = _forms_aktif(forms, status_map)
    # PERINGATAN: kartu di portal yg belum selesai tapi TAK ada di config (form
    # belum dipetakan) → tak akan terisi otomatis. Surface-kan agar tak diam2.
    tak_dikenal = _kartu_tak_dikenal(status_map, forms)
    if tak_dikenal:
        log("  ! FORM BELUM DIPETAKAN (ada di portal, TAK akan terisi otomatis): "
            + "; ".join(tak_dikenal))
    pilih_forms = None
    if args.forms:
        pilih_forms = [f.strip().lower() for f in args.forms.split(",") if f.strip()]
    for i, form in enumerate(forms, 1):
        if pilih_forms and not any(s in form["nama"].lower() for s in pilih_forms):
            continue
        sel = _form_selesai(status_map, form["nama"])   # None=kartu tak ada di detail
        if sel is None or form["nama"] not in aktif:
            # kartu tak ada, ATAU ada config lain lebih spesifik utk kartu yg sama
            n_skip_absen += 1                            # form tak berlaku utk peserta ini
            continue
        if args.resume and sel is True:
            n_skip_form += 1
            log(f"  ({i}/{len(forms)}) {form['nama'][:40]:40} SKIP (status: selesai diperiksa)")
            continue
        if not await dp._buka_form(page, fh, form["nama"]):
            log(f"  ({i}/{len(forms)}) LEWATI form {form['nama']!r}: kartu tak ketemu")
            tot_lewat += 1
            continue
        n_isi, n_lewat, n_skip, cat = await isi_satu_form(
            page, form, pvals, args.dry_run, log, resume=args.resume)
        tot_isi += n_isi
        tot_lewat += n_lewat
        tot_skip += n_skip
        tail = f"  ! {'; '.join(cat[:3])}" if cat else ""
        mode = " [DRY]" if args.dry_run else " [KIRIM]"
        if args.resume:
            mode += f" skip-terisi={n_skip}"
        log(f"  ({i}/{len(forms)}) {form['nama'][:40]:40} isi={n_isi} lewat={n_lewat}"
            f"{mode}{tail}")
        await _kembali_ke_detail(page, detail_url)

    # 5) selesaikan layanan → muncul dialog KUNCI → klik 'Konfirmasi' → catat waktu
    if args.selesaikan and not args.dry_run:
        # PENGAMAN: verifikasi form WAJIB sudah hijau/'Selesai diperiksa' SEBELUM
        # mengunci (jangan kunci data tak lengkap; lihat WAJIB_SELESAIKAN).
        status_now = await _peta_status(page)
        # ATURAN user: SEMUA Pemeriksaan Mandiri (tr) WAJIB hijau — tanpa pengecualian.
        belum_mandiri = await _mandiri_belum(page)
        mand_norm = {_norm_kartu(m) for m in belum_mandiri}
        # Kartu Nakes in-scope yg belum (LUAR_SCOPE boleh exempt — hanya utk Nakes/lab).
        # Buang yg sudah terhitung sbg Mandiri (cocokkan ternormalisasi).
        belum_nakes = [b for b in _kartu_belum_diluar_izin(status_now)
                       if _norm_kartu(b) not in mand_norm]
        belum = belum_mandiri + belum_nakes
        if belum and not args.paksa_selesai:
            log("  BATAL Selesaikan — masih ada form BELUM hijau. Data TIDAK dikunci:")
            for b in belum_mandiri:
                log(f"      - [MANDIRI wajib] {b}")
            for b in belum_nakes:
                log(f"      - [Nakes in-scope] {b}")
            log("    -> Lengkapi form di atas dulu (SEMUA Pemeriksaan Mandiri wajib hijau). "
                "Bila yakin boleh dikunci, jalankan ulang dgn --paksa-selesai.")
        elif await dp._klik_teks(page, fh, "Selesaikan Layanan"):
            if await _konfirmasi_kunci(page, log):
                res["selesai_ok"] = True
                res["waktu_selesai"] = _now()
                if not res["tgl_periksa"]:
                    res["tgl_periksa"] = tgl_iso or date.today().isoformat()
                log(f"  SELESAI PEMERIKSAAN (data DIKUNCI): waktu={res['waktu_selesai']}")
            else:
                log("  PERINGATAN: dialog 'Data Pemeriksaan akan Dikunci' / tombol "
                    "'Konfirmasi' tak ditemukan → data BELUM terkunci.")
        else:
            log("  PERINGATAN: 'Selesaikan Layanan' tak bisa diklik "
                "(syarat minimal belum lengkap? Pemeriksaan Mandiri/TB/BB/Gula/Tensi).")

    # tentukan status ringkas: SELESAI (selesaikan ok) / MULAI (baru dimulai) / DRAFT (dry)
    if args.dry_run:
        status = STATUS_LAYANAN_DRY
    elif res["selesai_ok"]:
        status = STATUS_LAYANAN_OK
    else:
        status = "MULAI"                      # diproses tapi belum 'Selesaikan Layanan'
    extra = f", form-selesai={n_skip_form}, field-terisi={tot_skip}" if args.resume else ""
    belum_map = f", BELUM-DIPETAKAN={len(tak_dikenal)}" if tak_dikenal else ""
    res["status"] = (f"{status} (isi={tot_isi}, lewat={tot_lewat}, "
                     f"tak-berlaku={n_skip_absen}{belum_map}{extra})")
    return res


def _tentukan_tabs(status, waktu_mulai, waktu_selesai):
    """Urutan tab listing dari kolom Excel (aturan user):
      - Waktu Mulai KOSONG                          → Belum Pemeriksaan (belum klik mulai)
      - Waktu Mulai terisi & Status Layanan=SELESAI → Selesai Pemeriksaan (sudah selesai)
      - Waktu Mulai terisi                          → Sedang Pemeriksaan (sudah dimulai)
    Tab lain tetap dicoba sbg fallback (portal = sumber kebenaran)."""
    s = str(status or "").strip().upper()
    selesai = bool(waktu_selesai) or s.startswith("SELESAI")
    if not waktu_mulai and not selesai:
        primary = "Belum Pemeriksaan"
    elif waktu_mulai and selesai:
        primary = "Selesai Pemeriksaan"
    elif selesai:                       # status SELESAI tapi waktu mulai kosong (mis. diproses manual)
        primary = "Selesai Pemeriksaan"
    else:
        primary = "Sedang Pemeriksaan"
    urut = ["Belum Pemeriksaan", "Sedang Pemeriksaan", "Selesai Pemeriksaan"]
    return [primary] + [t for t in urut if t != primary]


def _waktu_hadir_iso(p):
    v = getattr(p, "tanggal_pemeriksaan", None)
    return dp._tgl_iso(v) if v else None


def _peserta_values(p):
    """Kumpulkan nilai per-peserta utk form (Status Perkawinan + angka klinis)."""
    vals = {}
    sp = (getattr(p, "status_pernikahan", None) or "").strip().lower()
    if sp in PERNIKAHAN:
        vals[("perkawinan",)] = PERNIKAHAN[sp]
    # angka klinis diisi belakangan dari kolom Excel (lihat jalankan()).
    return vals


def _cari_kolom(ws, hdr, nama):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=hdr, column=c).value
        if v is not None and str(v).strip() == nama:
            return c
    return None


def _ensure_kolom(ws, hdr, nama):
    c = _cari_kolom(ws, hdr, nama)
    if c:
        return c
    c = ws.max_column + 1
    ws.cell(row=hdr, column=c, value=nama)
    return c


def _simpan(wb, path):
    try:
        wb.save(path)
        return True
    except PermissionError:
        log(f"GAGAL simpan '{path}': Excel sedang terbuka. TUTUP lalu jalankan lagi.")
        return False


async def jalankan(args):
    forms = baca_config(PEMETAAN)
    if not forms:
        raise SystemExit(f"Config form kosong di {PEMETAAN}. Jalankan buat_pemetaan_full.py dulu.")
    log(f"Config: {len(forms)} form, {sum(len(f['qs']) for f in forms)} pertanyaan.")

    ps = baca_excel(args.excel, KelompokUsia(args.kelompok), header_row=args.header_row)
    if not ps:
        raise SystemExit(f"Tidak ada data di {args.excel}.")
    wb = openpyxl.load_workbook(args.excel)
    ws = wb.worksheets[0]
    hdr = args.header_row + 1
    c_hadir = _cari_kolom(ws, hdr, KOL_STATUS_HADIR)
    c_waktu_hadir = _cari_kolom(ws, hdr, KOL_WAKTU_HADIR)
    c_status = _ensure_kolom(ws, hdr, KOL_STATUS_LAYANAN)
    c_waktu = _ensure_kolom(ws, hdr, KOL_WAKTU_LAYANAN)
    c_tgl = _ensure_kolom(ws, hdr, KOL_TGL_PERIKSA)
    c_mulai = _ensure_kolom(ws, hdr, KOL_WAKTU_MULAI)
    c_selesai = _ensure_kolom(ws, hdr, KOL_WAKTU_SELESAI)
    num_cols = _excel_num_cols(ws, hdr)
    if not _simpan(wb, args.excel):
        return 2

    mulai_idx = max(args.mulai - 1, 0)
    akhir_idx = len(ps) if args.jumlah <= 0 else min(mulai_idx + args.jumlah, len(ps))
    target = ps[mulai_idx:akhir_idx]
    if args.nik:
        target = [p for p in target if (p.nik or "") == args.nik]
        if not target:
            raise SystemExit(f"NIK {args.nik} tak ada di rentang itu.")
    log(f"Memproses {len(target)} peserta. Mode={'DRY-RUN' if args.dry_run else 'SUBMIT'}"
        f"{' +selesaikan' if args.selesaikan and not args.dry_run else ''}.")
    if num_cols:
        log(f"Kolom angka klinis terdeteksi: {list(num_cols)}")

    bot = CKGBot(headless=False, delay_ms=args.delay, cdp_url=args.cdp)
    try:
        await bot.connect_to_browser()
    except Exception as e:
        log(f"GAGAL connect Chrome: {e}. Pastikan port 9222 + login.")
        return 1
    page = bot._page
    fh = io.StringIO()                     # diag helpers butuh file handle utk out()

    n_ok = n_gagal = n_lewat = 0
    for p in target:
        row = p.baris_sumber
        label = f"baris {row} | {p.nama or '-'} | NIK={p.nik or '-'}"
        # hanya yg sudah HADIR (bila kolom hadir ada)
        if c_hadir is not None:
            sh = ws.cell(row=row, column=c_hadir).value
            if not (sh and str(sh).strip().startswith(STATUS_HADIR_TERMINAL)):
                log(f"LEWATI {label}: belum HADIR ({sh!r}).")
                n_lewat += 1
                continue
        # Status Layanan = SELESAI -> sudah tuntas, TAK PERLU di-search lagi (skip).
        # (dry-run tetap boleh memproses utk inspeksi.)
        st = ws.cell(row=row, column=c_status).value
        # Anti-dobel: status SELESAI → skip. TAPI saat --resume kita SENGAJA proses
        # ulang peserta SELESAI utk melengkapi form yg belum terisi (tak isi ulang
        # yg sudah hijau). (dry-run juga selalu boleh memproses utk inspeksi.)
        if (st and str(st).strip().startswith((STATUS_LAYANAN_OK, STATUS_SUDAH_SELESAI))
                and not args.dry_run and not args.resume):
            log(f"LEWATI {label}: Status Layanan sudah SELESAI ({st}).")
            n_lewat += 1
            continue

        # nilai angka klinis dari Excel utk peserta ini
        pvals_extra = {}
        for (fkey, sq), header in EXCEL_NUM.items():
            if header in num_cols:
                v = ws.cell(row=row, column=num_cols[header]).value
                if v not in (None, ""):
                    # cari nama form penuh yg memuat fkey
                    for f in forms:
                        if fkey in f["nama"]:
                            pvals_extra[("num", f["nama"], sq)] = v
        pvals = _peserta_values(p)
        pvals.update(pvals_extra)

        # tentukan tab kandidat (urutan) dari kolom status/waktu Excel.
        wmulai = ws.cell(row=row, column=c_mulai).value
        wselesai = ws.cell(row=row, column=c_selesai).value
        if args.tab:                       # override manual: pakai 1 tab itu saja
            tabs, auto_mulai = [args.tab], False
        else:
            tabs, auto_mulai = _tentukan_tabs(st, wmulai, wselesai), True
        # filter tanggal listing dari 'Waktu Hadir' (override --tanggal-filter bila ada)
        wh = ws.cell(row=row, column=c_waktu_hadir).value if c_waktu_hadir else None
        tgl_filter = args.tanggal_filter or dp._tgl_iso(wh)
        log(f"PROSES {label} ... (tab kandidat: {tabs}; tgl filter: {tgl_filter})")
        try:
            r = await proses_peserta(page, fh, bot, p, forms, args, pvals,
                                     tabs, auto_mulai, tgl_filter)
            status = r["status"]
            ws.cell(row=row, column=c_status, value=status)
            ws.cell(row=row, column=c_waktu, value=_now())
            if r["tgl_periksa"]:
                ws.cell(row=row, column=c_tgl, value=r["tgl_periksa"])
            if r["waktu_mulai"]:
                ws.cell(row=row, column=c_mulai, value=r["waktu_mulai"])
            if r["waktu_selesai"]:
                ws.cell(row=row, column=c_selesai, value=r["waktu_selesai"])
            log(f"  -> {status}")
            if status.startswith(STATUS_LAYANAN_OK) or status.startswith(STATUS_LAYANAN_DRY):
                n_ok += 1
            elif status.startswith(STATUS_SUDAH_SELESAI):
                n_lewat += 1            # sudah selesai di portal: dilewati, bukan gagal
            else:
                n_gagal += 1
        except Exception as e:
            pesan = f"GAGAL: {type(e).__name__}: {str(e)[:200]}"
            ws.cell(row=row, column=c_status, value=pesan)
            ws.cell(row=row, column=c_waktu, value=_now())
            log(f"  -> {pesan}")
            n_gagal += 1
            try:
                await page.goto(URL_PELAYANAN)
            except Exception:
                pass
        _simpan(wb, args.excel)

    await bot.stop()
    log("=" * 55)
    log(f"Selesai. OK={n_ok}  Gagal={n_gagal}  Dilewati={n_lewat}  "
        f"(kolom '{KOL_STATUS_LAYANAN}').")
    return 0


def main():
    ap = argparse.ArgumentParser(description="Otomasi pelayanan CKG (isi 22 form) via CDP.")
    ap.add_argument("--excel", required=True)
    ap.add_argument("--kelompok", default="lansia")
    ap.add_argument("--header-row", dest="header_row", type=int, default=0)
    ap.add_argument("--mulai", type=int, default=1)
    ap.add_argument("--jumlah", type=int, default=0)
    ap.add_argument("--nik", default="")
    ap.add_argument("--tab", default="",
                    help="Paksa 1 tab listing (Belum/Sedang/Selesai Pemeriksaan). "
                         "KOSONG (default) = AUTO per peserta dari kolom status/waktu Excel "
                         "(+ fallback cari di tab lain).")
    ap.add_argument("--aksi", default="Mulai")
    ap.add_argument("--mulai-pemeriksaan", dest="mulai_pemeriksaan", action="store_true",
                    help="Paksa klik 'Mulai Pemeriksaan' (hanya berlaku bila --tab dipaksa; "
                         "mode AUTO menentukannya dari tab tempat peserta ditemukan).")
    ap.add_argument("--tanggal-periksa", dest="tanggal_periksa", default="")
    ap.add_argument("--tanggal-filter", dest="tanggal_filter", default="",
                    help="Override tanggal filter listing (YYYY-MM-DD) utk SEMUA peserta. "
                         "Default: ikut 'Waktu Hadir' tiap baris Excel (HANYA dipakai di "
                         "jalur fallback filter Nama; jalur utama NIK tak butuh tanggal).")
    ap.add_argument("--filter-nama", dest="filter_nama", action="store_true",
                    help="Paksa jalur lama: filter 'Nama' + tanggal (lewati jalur cepat NIK). "
                         "Default: cari via NIK dulu (tanpa tanggal), Nama jadi fallback.")
    ap.add_argument("--forms", default="", help="Batasi form (substring, pisah koma).")
    ap.add_argument("--submit", dest="dry_run", action="store_false",
                    help="BENAR-BENAR kirim tiap form (default: dry-run, tak mengirim).")
    ap.add_argument("--dry-run", dest="dry_run", action="store_true", default=True)
    ap.add_argument("--selesaikan", action="store_true",
                    help="Klik 'Selesaikan Layanan' + 'Konfirmasi' di akhir (hanya bila "
                         "--submit). PERHATIAN: ini MENGUNCI data peserta (final, jadi "
                         "rapor yg dikirim ke WA/SATUSEHAT) & tak bisa diubah lagi.")
    ap.add_argument("--paksa-selesai", dest="paksa_selesai", action="store_true",
                    help="Tetap klik 'Selesaikan Layanan' walau ada form belum selesai "
                         "di luar out-of-scope (LEWATI pengaman). Pakai HANYA bila yakin "
                         "form sisa memang tak perlu diisi. MENGUNCI data (irreversible).")
    ap.add_argument("--resume", action="store_true",
                    help="Mode lanjut/resend: LEWATI pertanyaan yg sudah terisi (tak "
                         "ditimpa), isi hanya yg kosong, & tak Kirim ulang form yg utuh. "
                         "Memproses juga peserta ber-status SELESAI.")
    ap.add_argument("--delay", type=int, default=600)
    ap.add_argument("--cdp", default=S.CDP_URL)
    args = ap.parse_args()
    sys.exit(asyncio.run(jalankan(args)))


if __name__ == "__main__":
    main()
