"""
Buat TEMPLATE Excel untuk data pendaftaran REAL (produksi).

Menghasilkan `data/input/template_pendaftaran.xlsx` berisi baris header + 1 baris
contoh. ISI data asli Anda mulai baris ke-2, lalu hapus/timpa baris contoh.

Kolom mengikuti MAPPING di app/readers.py. Header WAJIB sama persis (huruf &
spasi) agar terbaca. Kolom WAJIB diisi untuk pendaftaran:
  NIK, Nama, Tanggal Lahir, Jenis Kelamin, No. WhatsApp
Kolom data pendukung boleh kosong -> dipakai nilai default (schema.py).

Jalankan:
    venv\\Scripts\\python.exe tools\\buat_template.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

OUT = os.path.join("data", "input", "template_pendaftaran.xlsx")

# (header, contoh, wajib?)
KOLOM = [
    ("NIK",               "3201234567890123", True),
    ("Nama",              "BUDI SANTOSO",     True),
    ("Tanggal Lahir",     "1990-01-15",       True),   # format ISO YYYY-MM-DD (anti-ambigu)
    ("Jenis Kelamin",     "L",                True),
    ("No. WhatsApp",      "81234567890",      True),
    ("No. HP",            "",                 False),
    ("Alamat",            "",                 False),
    ("Status Pernikahan", "Belum Kawin",      False),
    ("Disabilitas",       "Tidak memiliki disabilitas", False),
    ("Pekerjaan",         "Lainnya",          False),
    # Alamat Domisili = cascade wilayah. WAJIB & harus PERSIS sama dgn nama di
    # portal (mis. "Jawa Timur", "Kabupaten Gresik", "Driyorejo", nama desa).
    ("Provinsi",          "Jawa Timur",       True),
    ("Kabupaten/Kota",    "Kabupaten Gresik", True),
    ("Kecamatan",         "Driyorejo",        True),
    ("Kelurahan",         "Mojosarirejo",     True),
    ("Detail Alamat",     "Dsn. Contoh RT 01 RW 02", False),
    # --- Kolom HASIL (diisi otomatis oleh tools/jalankan_batch.py) ---
    # Baris yang 'No. Tiket'-nya sudah terisi akan DILEWATI saat dijalankan ulang.
    ("No. Tiket",         "",                 False),
    ("Status Daftar",     "",                 False),
    ("Waktu Daftar",      "",                 False),
]


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendaftaran"

    header_font = Font(bold=True, color="FFFFFF")
    fill_wajib = PatternFill("solid", fgColor="0F766E")   # teal - wajib
    fill_opsi = PatternFill("solid", fgColor="64748B")    # abu - opsional

    KOLOM_TEKS = ("NIK", "No. WhatsApp", "No. HP")
    for col, (judul, contoh, wajib) in enumerate(KOLOM, start=1):
        c = ws.cell(row=1, column=col, value=judul)
        c.font = header_font
        c.fill = fill_wajib if wajib else fill_opsi
        # baris contoh
        ws.cell(row=2, column=col, value=contoh)
        # NIK & nomor telepon: paksa format TEKS pada banyak baris agar 16 digit
        # tidak dibulatkan Excel (Excel hanya menyimpan 15 digit signifikan!).
        if judul in KOLOM_TEKS:
            for r in range(1, 501):
                ws.cell(row=r, column=col).number_format = "@"
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(
            16, len(judul) + 4)

    wb.save(OUT)
    print(f"Template dibuat: {OUT}")
    print("Kolom WAJIB (teal): NIK, Nama, Tanggal Lahir, Jenis Kelamin, No. WhatsApp")
    print("Kolom opsional (abu) boleh kosong -> pakai default.")
    print("Isi data asli mulai baris ke-2 (timpa baris contoh).")


if __name__ == "__main__":
    main()
