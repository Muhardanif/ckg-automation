"""
Tambah set form REMAJA / ANAK SEKOLAH (SMP-SMA) ke PEMETAAN_PELAYANAN_FULL.xlsx
(idempoten). Di-dump live 2026-06-16 dari peserta PUTRI ANA TASYA (perempuan, ~16 th).

Set remaja BERBEDA dari dewasa/lansia: kelompok ini punya form sendiri (Cemas/Depresi
Remaja, Kespro Putri, Malaria, Kelayakan Kebugaran, dst). Form yg NAMANYA mengandung
varian dewasa ('Perilaku Merokok' vs 'Perilaku Merokok - Anak Sekolah') ditangani
engine via pencocokan frag TERPANJANG (lihat _forms_aktif di pelayanan.py).

Default = baseline sehat: gejala/risiko -> Tidak; klinis -> Normal/Tidak ada.
Item ber-CEK perlu review user. Klinis angka (BB/TB/Sistol/Diastol/GDS) diisi dari
kolom Excel (lihat EXCEL_NUM di pelayanan.py); bila kolom kosong, field dilewati.

CATATAN cakupan: ini set REMAJA PEREMPUAN (SMP-SMA). Belum termasuk: 'Kesehatan
Reproduksi Putra' (remaja laki-laki) & kemungkinan varian usia SD (AZKA, ~12 th).

PAKAI (TUTUP Excel PEMETAAN dulu):
  venv\\Scripts\\python.exe tools\\tambah_form_remaja.py
"""
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

P = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                 "data", "output", "PEMETAAN_PELAYANAN_FULL.xlsx")
HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1565C0")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)

YT = "Ya  |  Tidak"

# nama_form, sumber, rows[(sq,tipe,judul,opsi,default,sel/cat? -> id, catatan)]
FORMS = [
    ("Faktor Risiko Malaria", "Default", [
        (100, "radio", "1. Gejala demam/sakit kepala/menggigil (malaria)?", YT, "Tidak", "sq_100i", ""),
        (101, "radio", "2. Pernah sakit malaria & obat tidak habis?", YT, "Tidak", "sq_101i", ""),
        (102, "radio", "3. Ada orang sakit malaria di tempat tinggal?", YT, "Tidak", "sq_102i", ""),
        (103, "radio", "4. Tinggal/riwayat dari daerah berisiko tinggi malaria?", YT, "Tidak", "sq_103i", ""),
    ]),
    ("Gejala Cemas Remaja", "Default", [
        (100, "radio", "1. (2 mgg) sering khawatir/tidak tenang/tegang?", YT, "Tidak", "sq_100i", ""),
        (101, "radio", "2. (2 mgg) berpikir berlebihan & tak terkendali?", YT, "Tidak", "sq_101i", ""),
        (102, "radio", "3. (2 mgg) sulit tidur & konsentrasi?", YT, "Tidak", "sq_102i", ""),
    ]),
    ("Gejala Depresi Remaja", "Default", [
        (100, "radio", "1. (2 mgg) sering sedih/tertekan tanpa sebab?", YT, "Tidak", "sq_100i", ""),
        (101, "radio", "2. (2 mgg) tak tertarik lagi pada kegiatan?", YT, "Tidak", "sq_101i", ""),
        (102, "radio", "3. (2 mgg) capek, sulit tidur, sulit fokus?", YT, "Tidak", "sq_102i", ""),
    ]),
    ("Kesehatan Reproduksi Putri - Anak Sekolah", "Default", [
        (100, "radio", "1. Apakah sudah mengalami menstruasi?", YT, "Ya", "sq_100i",
         "CEK: tergantung peserta (umumnya Ya utk remaja SMP-SMA)"),
        (101, "radio", "2. Pada usia berapa mengalami menstruasi pertama?",
         "<8 tahun atau >16 tahun  |  8 tahun -16 tahun", "8 tahun -16 tahun", "sq_101i",
         "KONDISIONAL (muncul bila sq100=Ya); default rentang normal 8-16 th"),
        (102, "radio", "3. Apakah pernah mengalami keputihan?", YT, "Tidak", "sq_102i", ""),
        (103, "radio", "4. Apakah pernah mengalami gatal di kemaluan?", YT, "Tidak", "sq_103i", ""),
    ]),
    ("Kuesioner Tingkat Aktivitas Fisik", "Default", [
        (100, "number", "1. (7 hari) berapa hari aktif fisik total >=60 menit?",
         "", "3", "sq_100i", "CEK: jumlah hari (0-7); 3 = baseline"),
        (101, "number", "2. Biasanya per minggu, berapa hari aktif fisik?",
         "", "3", "sq_101i", "CEK: jumlah hari (0-7); 3 = baseline"),
    ]),
    ("Kelayakan Tes Kebugaran", "Default", [
        (100, "radio", "1. Dokter pernah nyatakan masalah tulang/sendi?", YT, "Tidak", "sq_100i", ""),
        (101, "radio", "2. Dokter pernah nyatakan masalah jantung?", YT, "Tidak", "sq_101i", ""),
        (102, "radio", "3. Asma / pernah asma saat latihan?", YT, "Tidak", "sq_102i", ""),
        (103, "radio", "4. Pernah kehilangan kesadaran/pingsan?", YT, "Tidak", "sq_103i", ""),
    ]),
    ("Perilaku Merokok - Anak Sekolah", "Default", [
        (100, "radio", "1. Apakah Anda merokok dalam setahun terakhir?", YT, "Tidak", "sq_100i", ""),
        (104, "radio", "2. Apakah Anda terpapar asap rokok orang lain?", YT, "Tidak", "sq_104i", ""),
    ]),
    ("Faktor Risiko Hepatitis SMP dan SMA", "Default", [
        (100, "radio", "1. Pernah tes Hepatitis B hasil positif?", YT, "Tidak", "sq_100i", ""),
        (101, "radio", "2. Ibu/saudara kandung menderita Hepatitis B/C?", YT, "Tidak", "sq_101i", ""),
        (102, "radio", "3. Pernah hubungan seksual berisiko/tanpa pengaman?", YT, "Tidak", "sq_102i", ""),
        (103, "radio", "4. Pernah menerima transfusi darah?", YT, "Tidak", "sq_103i", ""),
        (104, "radio", "5. Pernah cuci darah/hemodialisis?", YT, "Tidak", "sq_104i", ""),
        (105, "radio", "6. Pernah pakai narkoba/zat adiktif?", YT, "Tidak", "sq_105i", ""),
        (106, "radio", "7. Apakah Anda ODHIV?", YT, "Tidak", "sq_106i", ""),
        (107, "radio", "8. Pernah pengobatan Hepatitis C & tidak sembuh?", YT, "Tidak", "sq_107i", ""),
    ]),
    ("Skrining Telinga dan Mata - Anak Sekolah", "Default", [
        (100, "radio", "1. Gangguan pendengaran telinga kanan?",
         "Normal  |  Ada indikasi gangguan pendengaran", "Normal", "sq_100i", ""),
        (101, "radio", "2. Gangguan pendengaran telinga kiri?",
         "Normal  |  Ada indikasi gangguan pendengaran", "Normal", "sq_101i", ""),
        (102, "radio", "3. Serumen impaksi telinga kanan?",
         "Tidak ada serumen impaksi  |  Ada serumen impaksi", "Tidak ada serumen impaksi", "sq_102i", ""),
        (103, "radio", "4. Serumen impaksi telinga kiri?",
         "Tidak ada serumen impaksi  |  Ada serumen impaksi", "Tidak ada serumen impaksi", "sq_103i", ""),
        (104, "radio", "5. Infeksi telinga kanan?",
         "Tidak ada infeksi telinga  |  Ada infeksi telinga", "Tidak ada infeksi telinga", "sq_104i", ""),
        (105, "radio", "6. Infeksi telinga kiri?",
         "Tidak ada infeksi telinga  |  Ada infeksi telinga", "Tidak ada infeksi telinga", "sq_105i", ""),
        (106, "radio", "7. Mata kanan (luar)?",
         "Normal  |  Curiga kelainan mata", "Normal", "sq_106i", ""),
        (107, "radio", "8. Mata kiri (luar)?",
         "Normal  |  Curiga kelainan mata", "Normal", "sq_107i", ""),
        (108, "radio", "9. Tajam penglihatan mata kanan?",
         "Normal (visus 6/6 - 6/9)  |  Ada indikasi gangguan penglihatan (visus <6/9)",
         "Normal (visus 6/6 - 6/9)", "sq_108i", ""),
        (109, "radio", "10. Tajam penglihatan mata kiri?",
         "Normal (visus 6/6 - 6/9)  |  Ada indikasi gangguan penglihatan (visus <6/9)",
         "Normal (visus 6/6 - 6/9)", "sq_109i", ""),
        (110, "radio", "11. Perlu rujukan?", "Tidak  |  Ya", "Tidak", "sq_110i",
         "CEK: polaritas (opsi: Tidak | Ya)"),
    ]),
    ("Pemeriksaan Gigi - Anak", "Default", [
        (100, "radio", "1. Berapa jumlah gigi karies?",
         "Tidak ada  |  1  |  2  |  3  |  >3", "Tidak ada", "sq_100i", ""),
    ]),
    ("Gizi Anak Sekolah", "Excel", [
        (100, "number", "1. Berat Badan", "", "", "sq_100i", "dari kolom Excel 'Berat Badan'"),
        (101, "number", "2. Tinggi Badan", "", "", "sq_101i", "dari kolom Excel 'Tinggi Badan'"),
    ]),
    ("Tekanan Darah Anak dan Remaja", "Excel", [
        (100, "number", "1. Tekanan Darah Sistol", "", "", "sq_100i", "dari kolom Excel 'Sistolik'"),
        (101, "number", "2. Tekanan Darah Diastol", "", "", "sq_101i", "dari kolom Excel 'Diastolik'"),
    ]),
    ("Pemeriksaan Gula Darah Remaja", "Excel/Default", [
        (100, "radio", "1. Pernah dinyatakan diabetes oleh dokter?", YT, "Tidak", "sq_100i", ""),
        (102, "number", "2. Gula Darah Sewaktu (GDS)", "", "", "sq_102i", "dari kolom Excel 'GDS'"),
        (103, "number", "3. Gula Darah Sewaktu Kedua (GDS 2)", "", "", "sq_103i",
         "KONDISIONAL & OPSIONAL: muncul jika GDS 1 prediabetes; dari kolom Excel 'GDS 2'"),
    ]),
]


def main():
    wb = openpyxl.load_workbook(P)
    ov = wb["Daftar Form"]
    tambah = 0
    for nama, sumber, rows in FORMS:
        sheet = nama[:31]
        if sheet in wb.sheetnames:
            print(f"  sheet '{sheet}' sudah ada, lewati.")
            continue
        ws = wb.create_sheet(sheet)
        cols = ["No", "sq", "Tipe", "Pertanyaan", "Opsi", "Nilai Default (ISI)",
                "id base / PPM", "Catatan"]
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = HDR
            cell.fill = FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for i, (sq, tipe, judul, opsi, default, sid, cat) in enumerate(rows, 1):
            ws.append([i, sq, tipe, judul, opsi, default, sid, cat])
        for k, w in enumerate([5, 7, 10, 44, 50, 24, 14, 34], 1):
            ws.column_dimensions[chr(64 + k)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = WRAP
                cell.border = THIN
        ov.append([ov.max_row, nama, "Ya", sumber, len(rows), ws.title])
        print(f"  + ditambah: {nama} ({len(rows)} pertanyaan)")
        tambah += 1
    try:
        wb.save(P)
    except PermissionError:
        raise SystemExit("GAGAL simpan: TUTUP Excel PEMETAAN_PELAYANAN_FULL.xlsx lalu ulangi.")
    print(f"[OK] {tambah} form remaja ditambahkan ke {P}")


if __name__ == "__main__":
    main()
