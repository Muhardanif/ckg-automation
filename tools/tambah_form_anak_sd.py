"""
Tambah form khusus ANAK SEKOLAH SD (usia 7-12) ke PEMETAAN_PELAYANAN_FULL.xlsx
(idempoten). Di-dump live 2026-06-16 dari peserta AZKA (laki-laki, ~12 th).

Anak SD berbagi MAYORITAS form dgn remaja (Malaria, Cemas/Depresi, Aktivitas Fisik,
Kelayakan Kebugaran, Merokok, Gizi, Tensi, Telinga-Mata, Gigi → reuse config remaja).
Yg BERBEDA & ditambah di sini hanya 4 form:
  - Faktor Risiko Gula Darah Anak (1r)
  - Kesehatan Reproduksi Putra - Anak Sekolah (3r) [versi laki-laki, beda dari Putri]
  - Faktor Risiko Hepatitis SD (4r) [subset dari Hepatitis SMP/SMA]
  - Pemeriksaan Gula Darah Anak (diabetes r + GDS num) [beda nama dari versi Remaja]

Default = baseline sehat (risiko/gejala → Tidak). Engine pakai pencocokan frag
TERPANJANG (_forms_aktif) shg form 'Putra' vs 'Putri' / 'Hepatitis SD' vs 'SMP-SMA'
/ 'Gula Darah Anak' vs 'Remaja' tidak saling tabrak.

PAKAI (TUTUP Excel PEMETAAN dulu):
  venv\\Scripts\\python.exe tools\\tambah_form_anak_sd.py
"""
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

P = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                 "data", "output", "PEMETAAN_PELAYANAN_FULL.xlsx")
HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="00838F")
WRAP = Alignment(vertical="top", wrap_text=True)
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)
YT = "Ya  |  Tidak"

FORMS = [
    ("Faktor Risiko Gula Darah Anak", "Default", [
        (100, "radio", "1. Pernah dinyatakan diabetes/kencing manis oleh dokter?",
         YT, "Tidak", "sq_100i", ""),
        # sq_102-105 KONDISIONAL (muncul setelah sq_100). Perhatikan opsi: sebagian
        # 'Iya' (bukan 'Ya'). Semua default baseline = Tidak (tanpa gejala/risiko).
        (102, "radio", "2. Sering merasa sangat lapar & makan lebih banyak?",
         "Iya  |  Tidak", "Tidak", "sq_102i", "KONDISIONAL"),
        (103, "radio", "3. Sering merasa haus meskipun sudah banyak minum?",
         "Ya  |  Tidak", "Tidak", "sq_103i", "KONDISIONAL"),
        (104, "radio", "4. Tetap turun berat badan meskipun makan banyak?",
         "Iya  |  Tidak", "Tidak", "sq_104i", "KONDISIONAL"),
        (105, "radio", "5. Ada keluarga (saudara kandung) yang diabetes?",
         "Iya  |  Tidak", "Tidak", "sq_105i", "KONDISIONAL"),
    ]),
    ("Kesehatan Reproduksi Putra - Anak Sekolah", "Default", [
        (100, "radio", "1. Gatal di kemaluan / kencing keruh?", YT, "Tidak", "sq_100i", ""),
        (101, "radio", "2. Nyeri/tidak nyaman saat BAK atau BAB?", YT, "Tidak", "sq_101i", ""),
        (102, "radio", "3. Ada luka di anus atau dubur?", YT, "Tidak", "sq_102i", ""),
    ]),
    ("Faktor Risiko Hepatitis SD", "Default", [
        (100, "radio", "1. Pernah tes Hepatitis B hasil positif?", YT, "Tidak", "sq_100i", ""),
        (101, "radio", "2. Ibu/saudara kandung menderita Hepatitis B?", YT, "Tidak", "sq_101i", ""),
        (102, "radio", "3. Pernah menerima transfusi darah?", YT, "Tidak", "sq_102i", ""),
        (103, "radio", "4. Pernah cuci darah/hemodialisis?", YT, "Tidak", "sq_103i", ""),
    ]),
    ("Pemeriksaan Gula Darah Anak", "Excel/Default", [
        (100, "radio", "1. Pernah dinyatakan diabetes oleh dokter?", YT, "Tidak", "sq_100i", ""),
        (102, "number", "2. Gula Darah Sewaktu (GDS)", "", "", "sq_102i", "dari kolom Excel 'GDS'"),
        (103, "number", "3. Gula Darah Sewaktu Kedua (GDS 2)", "", "", "sq_103i",
         "KONDISIONAL & OPSIONAL: muncul jika GDS 1 prediabetes; dari kolom Excel 'GDS 2'"),
    ]),
]


def main():
    wb = openpyxl.load_workbook(P)
    ov = wb["Daftar Form"]
    tambah = 0
    for nama, sumber, rows in FORMS:
        sheet = nama[:31]
        if sheet in wb.sheetnames:
            print(f"  sheet '{sheet}' sudah ada, lewati.")
            continue
        ws = wb.create_sheet(sheet)
        cols = ["No", "sq", "Tipe", "Pertanyaan", "Opsi", "Nilai Default (ISI)",
                "id base / PPM", "Catatan"]
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = HDR
            cell.fill = FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for i, (sq, tipe, judul, opsi, default, sid, cat) in enumerate(rows, 1):
            ws.append([i, sq, tipe, judul, opsi, default, sid, cat])
        for k, w in enumerate([5, 7, 10, 44, 50, 24, 14, 34], 1):
            ws.column_dimensions[chr(64 + k)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = WRAP
                cell.border = THIN
        ov.append([ov.max_row, nama, "Ya", sumber, len(rows), ws.title])
        print(f"  + ditambah: {nama} ({len(rows)} pertanyaan)")
        tambah += 1
    try:
        wb.save(P)
    except PermissionError:
        raise SystemExit("GAGAL simpan: TUTUP Excel PEMETAAN_PELAYANAN_FULL.xlsx lalu ulangi.")
    print(f"[OK] {tambah} form anak-SD ditambahkan ke {P}")


if __name__ == "__main__":
    main()
