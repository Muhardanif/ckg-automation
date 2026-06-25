"""
KONFIRMASI HADIR: konfirmasi kehadiran peserta yang SUDAH terdaftar (mode CDP).

Membaca Excel hasil pendaftaran, lalu untuk tiap baris ber-Status Daftar
'SUKSES' membuka halaman /ckg-pendaftaran-individu, mencari peserta via NIK,
dan meng-klik 'Konfirmasi Hadir'.

Fitur penting (mirip tools/jalankan_batch.py):
- HANYA memproses baris 'Status Daftar' = SUKSES (yang gagal/terminal dilewati).
- RESUMABLE / anti-dobel: hasil ditulis ke kolom 'Status Hadir' / 'Waktu Hadir'.
  Baris yang 'Status Hadir'-nya sudah terisi terminal (HADIR / SUDAH HADIR)
  otomatis DILEWATI saat dijalankan ulang.

Filter tanggal diatur OTOMATIS per baris sesuai 'Waktu Daftar' di Excel (tabel
portal difilter per Tanggal Pemeriksaan; tanggal daftar = tanggal pemeriksaan).
Pakai --tanggal YYYY-MM-DD untuk meng-override semua baris ke satu tanggal.

PERSIAPAN (sama seperti pendaftaran):
  1. Jalankan Chrome dengan --remote-debugging-port=9222 (1_mulai_chrome.bat),
     login manual ke portal, buka menu CKG Umum > Cari/Daftarkan Individu.
  2. TUTUP file Excel di aplikasi Excel (skrip menulis balik ke file itu).

PAKAI:
  venv\\Scripts\\python.exe tools\\konfirmasi_hadir.py --excel data/input/template_pendaftaran.xlsx
  # opsi: --mulai 1 --jumlah 50 --kelompok dewasa
  # uji 1 NIK saja (abaikan Excel utk loop, tetap tulis balik bila ketemu di Excel):
  venv\\Scripts\\python.exe tools\\konfirmasi_hadir.py --excel ... --nik 3515xxxxxxxxxxxx
"""
import argparse
import asyncio
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl                                                  # noqa: E402

from app.automation.ckg_bot import CKGBot, LewatiPesertaError    # noqa: E402
from app.automation import selectors as S                        # noqa: E402
from app.schema import KelompokUsia                              # noqa: E402
from app.readers import baca_excel, validasi                     # noqa: E402
from app.excel_hasil import (KOL_STATUS, KOL_WAKTU,              # noqa: E402
                             KOL_STATUS_HADIR, KOL_WAKTU_HADIR,
                             STATUS_HADIR, STATUS_HADIR_TERMINAL)

# Hanya baris dengan Status Daftar diawali ini yang dikonfirmasi hadir.
STATUS_DAFTAR_SUKSES = "SUKSES"


def log(msg):
    print(f"[HADIR] {msg}", flush=True)


def _tgl_iso(val):
    """Ambil tanggal ISO 'YYYY-MM-DD' dari sel 'Waktu Daftar'.
    Menerima datetime, atau string ISO ('2026-06-12T10:30:00') / 'YYYY-MM-DD ...'.
    Kembalikan None bila tak bisa diuraikan."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return s[:10]
    return None


def _ensure_kolom(ws, header_row_1based, nama):
    """Index kolom (1-based) dgn header `nama`; buat baru bila belum ada."""
    last_col = ws.max_column
    for c in range(1, last_col + 1):
        v = ws.cell(row=header_row_1based, column=c).value
        if v is not None and str(v).strip() == nama:
            return c
    col = last_col + 1
    ws.cell(row=header_row_1based, column=col, value=nama)
    return col


def _cari_kolom(ws, header_row_1based, nama):
    """Index kolom (1-based) dgn header `nama`, atau None bila tak ada."""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row_1based, column=c).value
        if v is not None and str(v).strip() == nama:
            return c
    return None


def _simpan(wb, path):
    try:
        wb.save(path)
        return True
    except PermissionError:
        log(f"GAGAL menyimpan '{path}': file sedang dibuka di Excel. "
            f"TUTUP Excel lalu jalankan lagi (baris yang sudah hadir tidak "
            f"akan diulang).")
        return False


async def jalankan(args):
    # 1) baca data ternormalisasi (Peserta) + buka workbook utk tulis-balik
    ps = baca_excel(args.excel, KelompokUsia(args.kelompok),
                    header_row=args.header_row)
    if not ps:
        raise SystemExit(f"Tidak ada data terbaca di {args.excel}.")

    wb = openpyxl.load_workbook(args.excel)
    ws = wb.worksheets[0]
    hdr = args.header_row + 1
    c_status_daftar = _cari_kolom(ws, hdr, KOL_STATUS)   # 'Status Daftar'
    if c_status_daftar is None:
        raise SystemExit(
            f"Kolom '{KOL_STATUS}' tak ada di Excel. Jalankan pendaftaran dulu "
            f"(kolom itu dibuat oleh tools/jalankan_batch.py).")
    c_waktu_daftar = _cari_kolom(ws, hdr, KOL_WAKTU)     # 'Waktu Daftar' (utk tgl filter)
    c_hadir = _ensure_kolom(ws, hdr, KOL_STATUS_HADIR)
    c_waktu = _ensure_kolom(ws, hdr, KOL_WAKTU_HADIR)
    if not _simpan(wb, args.excel):     # gagal cepat bila Excel masih terbuka
        return 2

    # 2) tentukan rentang baris
    mulai_idx = max(args.mulai - 1, 0)
    akhir_idx = len(ps) if args.jumlah <= 0 else min(mulai_idx + args.jumlah, len(ps))
    target = ps[mulai_idx:akhir_idx]
    if args.nik:
        target = [p for p in target if (p.nik or "") == args.nik]
        if not target:
            raise SystemExit(f"NIK {args.nik} tak ada di rentang baris itu.")
    log(f"Total {len(ps)} baris; memproses {len(target)} "
        f"(baris data {mulai_idx + 1}..{akhir_idx}"
        f"{'; filter NIK '+args.nik if args.nik else ''}).")

    # 3) sambung ke Chrome (sekali)
    bot = CKGBot(headless=False, delay_ms=args.delay, cdp_url=args.cdp)
    try:
        await bot.connect_to_browser()
    except Exception as e:
        log(f"GAGAL connect ke Chrome: {type(e).__name__}: {e}")
        log("Pastikan Chrome jalan dgn --remote-debugging-port=9222 & sudah login.")
        return 1

    n_hadir = n_gagal = n_lewat = 0
    for p in target:
        row = p.baris_sumber
        label = f"baris {row} | NIK={p.nik or '-'} | {p.nama or '-'}"

        # a) hanya proses yang Status Daftar = SUKSES
        sd = ws.cell(row=row, column=c_status_daftar).value
        if sd is None or not str(sd).strip().startswith(STATUS_DAFTAR_SUKSES):
            log(f"LEWATI {label}: Status Daftar bukan SUKSES ({sd!r}).")
            n_lewat += 1
            continue

        # a2) tentukan tanggal filter = tanggal daftar peserta (dari 'Waktu
        #     Daftar'), kecuali di-override --tanggal. Tabel difilter per tanggal,
        #     jadi ini wajib agar baris peserta muncul saat dicari via NIK.
        if args.tanggal:
            tgl_filter = args.tanggal
        elif c_waktu_daftar is not None:
            tgl_filter = _tgl_iso(ws.cell(row=row, column=c_waktu_daftar).value)
        else:
            tgl_filter = None

        # b) sudah dikonfirmasi hadir? (anti-dobel) -> lewati
        sh = ws.cell(row=row, column=c_hadir).value
        if sh is not None and str(sh).strip().startswith(STATUS_HADIR_TERMINAL):
            log(f"LEWATI {label}: sudah {str(sh).strip()}.")
            n_lewat += 1
            continue

        # c) pra-cek NIK valid (16 digit) - tanpa NIK tak bisa cari
        err = [e for e in validasi(p) if "NIK" in e]
        if err:
            pesan = "ERROR data: " + "; ".join(err)
            log(f"LEWATI {label}: {pesan}")
            ws.cell(row=row, column=c_hadir, value=pesan)
            ws.cell(row=row, column=c_waktu,
                    value=datetime.now().isoformat(timespec="seconds"))
            _simpan(wb, args.excel)
            n_lewat += 1
            continue

        # d) konfirmasi hadir
        log(f"PROSES {label} ...")
        try:
            await bot.konfirmasi_hadir_satu(
                p, on_step=lambda nama, info="": None,
                tanggal_filter=tgl_filter)
            ws.cell(row=row, column=c_hadir, value=STATUS_HADIR)
            ws.cell(row=row, column=c_waktu,
                    value=datetime.now().isoformat(timespec="seconds"))
            log(f"HADIR {label}")
            n_hadir += 1
        except LewatiPesertaError as e:
            # SUDAH HADIR / TIDAK DITEMUKAN -> tandai terminal, lanjut.
            pesan = f"{e.status_prefix}: {str(e)[:200]}"
            ws.cell(row=row, column=c_hadir, value=pesan)
            ws.cell(row=row, column=c_waktu,
                    value=datetime.now().isoformat(timespec="seconds"))
            log(f"LEWATI {label}: {pesan}")
            n_lewat += 1
            try:
                await bot._page.reload()
                await bot._page.wait_for_load_state("networkidle")
            except Exception:
                pass
        except Exception as e:
            pesan = f"GAGAL: {str(e)[:300]}"
            ws.cell(row=row, column=c_hadir, value=pesan)
            ws.cell(row=row, column=c_waktu,
                    value=datetime.now().isoformat(timespec="seconds"))
            log(f"GAGAL {label}: {type(e).__name__}: {str(e)[:200]}")
            n_gagal += 1
            try:
                await bot._page.reload()
                await bot._page.wait_for_load_state("networkidle")
            except Exception:
                pass
        _simpan(wb, args.excel)

    await bot.stop()
    log("=" * 55)
    log(f"Selesai. Hadir={n_hadir}  Gagal={n_gagal}  Dilewati={n_lewat}")
    log(f"Hasil ditulis balik ke: {args.excel} "
        f"(kolom '{KOL_STATUS_HADIR}' / '{KOL_WAKTU_HADIR}').")
    log("=" * 55)
    return 0


def main():
    ap = argparse.ArgumentParser(
        description="Konfirmasi hadir peserta CKG dari Excel (CDP).")
    ap.add_argument("--excel", required=True,
                    help="Path file Excel (hasil pendaftaran; hasil hadir ditulis "
                         "balik ke file ini).")
    ap.add_argument("--kelompok", default="dewasa",
                    help="bayi/balita/dewasa/lansia (default dewasa).")
    ap.add_argument("--header-row", dest="header_row", type=int, default=0,
                    help="Indeks baris header (0 = baris pertama).")
    ap.add_argument("--mulai", type=int, default=1,
                    help="Mulai dari baris data ke-berapa (1-based). Default 1.")
    ap.add_argument("--jumlah", type=int, default=0,
                    help="Berapa baris diproses (0 = semua sampai akhir).")
    ap.add_argument("--nik", default="",
                    help="Konfirmasi hanya NIK ini (untuk uji 1 peserta).")
    ap.add_argument("--tanggal", default="",
                    help="Override tanggal filter (YYYY-MM-DD) utk SEMUA baris. "
                         "Default: ikut 'Waktu Daftar' tiap baris di Excel.")
    ap.add_argument("--delay", type=int, default=800,
                    help="Jeda antar-aksi (ms). Naikkan bila koneksi lambat.")
    ap.add_argument("--cdp", default=S.CDP_URL,
                    help=f"URL Chrome remote-debugging (default {S.CDP_URL}).")
    args = ap.parse_args()
    sys.exit(asyncio.run(jalankan(args)))


if __name__ == "__main__":
    main()
